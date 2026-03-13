#!/usr/bin/env python3
"""
build.py — Core data processing and map/page generation for CHASERMap.

This script has two roles:

1.  **Data pipeline library** — exposes pure functions (load_races,
    load_statewide_races, shapefile_to_geojson, load_places, build_city_map,
    and all _fmt_* helpers) that are imported by ingest.py and app.py.
    These functions are safe to import because no side-effectful code runs at
    module level — everything is guarded by `if __name__ == "__main__"`.

2.  **Static HTML generator** — when run directly (`python3 build.py`) it
    executes main(), which:
      a. Reads data/tracer_2026_all_districts.csv (scraped by scraper.py)
      b. Downloads / caches Colorado district shapefiles from the US Census
      c. Generates map/index.html — a fully self-contained Leaflet map with
         all candidate data embedded as JSON inside <script> tags.
      d. Generates static race and candidate HTML pages (now superseded by the
         Flask app, but kept for reference / static hosting fallback).

Data flow:
    scraper.py  →  data/tracer_2026_all_districts.csv
                →  data/tracer_2026_statewide.csv
    build.py    →  data/co_senate_districts.json   (cached shapefile)
                →  data/co_house_districts.json    (cached shapefile)
                →  data/co_places.json             (cached place centroids)
                →  map/index.html                  (Leaflet district map)
    ingest.py   →  data/chaser.db                  (SQLite for Flask)

Usage:
    python3 build.py          # regenerate map/index.html from latest CSV data
"""

import csv
import io
import json
import re
import urllib.request
import zipfile
from pathlib import Path
import shapefile  # pyshp — pure-Python shapefile reader (no GDAL required)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
DATA_DIR = Path(__file__).parent / "data"
MAP_DIR  = Path(__file__).parent / "map"

# TRACER export CSVs produced by scraper.py
CSV_FILE         = DATA_DIR / "tracer_2026_all_districts.csv"
STATEWIDE_CSV    = DATA_DIR / "tracer_2026_statewide.csv"

# Output: the self-contained Leaflet map HTML
OUTPUT_HTML      = MAP_DIR  / "index.html"

# US Census Bureau cartographic boundary shapefiles for Colorado (2022 vintage).
# These are ~500k-resolution (simplified) boundaries, suitable for web display.
# Downloaded once and cached as GeoJSON to avoid repeated HTTP requests.
SHAPEFILE_URLS = {
    "Senate": "https://www2.census.gov/geo/tiger/GENZ2022/shp/cb_2022_08_sldu_500k.zip",
    "House":  "https://www2.census.gov/geo/tiger/GENZ2022/shp/cb_2022_08_sldl_500k.zip",
}
# Local cache paths for the converted GeoJSON files
GEOJSON_CACHE = {
    "Senate": DATA_DIR / "co_senate_districts.json",
    "House":  DATA_DIR / "co_house_districts.json",
}
# Attribute field in the shapefile's DBF that contains the district number.
# SLDUST = State Legislative District Upper (Senate), SLDLST = Lower (House).
DISTRICT_FIELD = {
    "Senate": "SLDUST",
    "House":  "SLDLST",
}

# Colorado incorporated places shapefile from TIGER/Line (for city-to-district mapping)
PLACES_URL        = "https://www2.census.gov/geo/tiger/TIGER2022/PLACE/tl_2022_08_place.zip"
PLACES_CACHE      = DATA_DIR / "co_places.json"     # cached centroid lookup

# Official CO legislature roster (Excel) — used for incumbent detection.
# Download from: https://leg.colorado.gov/legislators
LEGISLATORS_FILE  = DATA_DIR / "co_legislators.xlsx"

# ---------------------------------------------------------------------------
# Incumbent detection
# ---------------------------------------------------------------------------

# Regex to strip generational suffixes before name comparison.
# Handles "JR", "JR.", "SR", "SR.", "II", "III", "IV", "V" (case-insensitive).
_SUFFIX_RE = re.compile(r'\b(JR\.?|SR\.?|II|III|IV|V)\b\.?', re.IGNORECASE)

# Some incumbents file campaign finance reports under a preferred name that
# differs from the official roster name.  These overrides map (chamber, district)
# to the first name *as it appears in TRACER*, so the match logic uses the
# TRACER name directly rather than the legislature's name.
_NICKNAME_FIRST_NAMES: dict[tuple, str] = {
    ("Senate", "11"): "THOMAS",   # Tony Exum Sr. files as Thomas E. Exum Sr.
    ("House",  "40"): "NIKKI",    # Naquetta Ricks files as Nikki Ricks
}

# Legislators who are not present in the official Excel file (e.g. mid-session
# appointments or corrections) but should still be flagged as incumbents.
# Keyed by (chamber, district), value is the last name in TRACER (upper-case,
# suffix already stripped).
_EXTRA_INCUMBENTS: dict[tuple, str] = {
    ("House", "33"): "NGUYEN",    # Kenny Nguyen (HD-33)
}


def load_incumbents() -> dict:
    """Load the current Colorado legislators from the official Excel roster.

    The workbook has two sheets — "Representatives" (House) and "Senators"
    (Senate).  Data rows start at row 3 (row 1 = header, row 2 = blank).
    Expected column layout (0-indexed): ... [1]=first, [2]=last, [6]=district.

    Returns:
        dict mapping (chamber, district_str) → (last_upper, first_upper)
        e.g. {("Senate", "3"): ("BRIDGES", "JEFF"), ...}
        Returns {} if openpyxl is not installed or the file is missing.
    """
    try:
        import openpyxl
    except ImportError:
        print("  \u26a0  openpyxl not installed — no incumbent flags (pip install openpyxl)")
        return {}
    if not LEGISLATORS_FILE.exists():
        print(f"  \u26a0  {LEGISLATORS_FILE.name} not found — no incumbent flags")
        return {}

    wb = openpyxl.load_workbook(LEGISLATORS_FILE)
    result = {}
    for sheet, chamber in [("Representatives", "House"), ("Senators", "Senate")]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[1]:  # skip blank rows
                continue
            first = row[1].upper().strip()
            last  = row[2].upper().strip()
            dist  = str(int(row[6]))  # normalize "03" → "3"
            result[(chamber, dist)] = (last, first)
    print(f"  Loaded {len(result)} incumbents from {LEGISLATORS_FILE.name}")
    return result


def is_incumbent(name: str, chamber: str, dist: str, incumbents: dict) -> bool:
    """Return True if the TRACER candidate name matches the sitting incumbent.

    TRACER names are stored as "LAST, FIRST MIDDLE" in all-caps.  The
    official roster uses a different format and may omit suffixes.  This
    function handles several real-world mismatches:

      - Generational suffixes  — stripped via _SUFFIX_RE before comparison
      - Compound last names    — e.g. TRACER "WILSON" matches roster "ZAMORA WILSON"
      - Middle-name-as-first   — e.g. TRACER "TIMOTHY JARVIS" where roster has "JARVIS"
      - Short-form first names — e.g. "RODNEY" matches "ROD" via 3-char prefix check
      - Explicit nickname map  — _NICKNAME_FIRST_NAMES overrides per (chamber, dist)
      - Extra incumbents       — _EXTRA_INCUMBENTS handles omissions in the Excel file

    Args:
        name:       Candidate name from TRACER CSV, format "LAST, FIRST [MIDDLE]".
        chamber:    "Senate" or "House".
        dist:       District number string, e.g. "3".
        incumbents: Dict returned by load_incumbents().

    Returns:
        True if the candidate is the current incumbent for this seat.
    """
    info = incumbents.get((chamber, dist))

    # Check _EXTRA_INCUMBENTS first — these are last-name-only matches for
    # legislators not present in the Excel file.
    extra_last = _EXTRA_INCUMBENTS.get((chamber, dist))

    # Parse TRACER name: "LAST, FIRST MIDDLE [JR.]" → separate components
    parts = name.strip().upper().split(",", 1)
    raw_last        = parts[0].strip()
    raw_first_field = parts[1].strip() if len(parts) > 1 else ""
    tracer_last     = _SUFFIX_RE.sub("", raw_last).strip()   # strip Jr./Sr./II/etc.
    tracer_first    = raw_first_field.split()[0] if raw_first_field else ""

    # Extra-incumbent override: last-name check only
    if extra_last and tracer_last == extra_last:
        return True

    if not info:
        return False  # no incumbent on record for this seat
    leg_last, leg_first = info

    # Last-name match — also accepts TRACER last being one word of a compound
    # roster name (e.g. TRACER "WILSON" matches roster "ZAMORA WILSON")
    last_match = tracer_last == leg_last or tracer_last in leg_last.split()
    if not last_match:
        return False

    # Resolve effective first name: use the nickname override if one exists,
    # otherwise fall back to the legislator roster first name.
    eff_first = _NICKNAME_FIRST_NAMES.get((chamber, dist), leg_first)

    return (
        tracer_first == eff_first or                # exact first-name match
        tracer_first.startswith(eff_first[:3]) or   # prefix match (ROD/RODNEY)
        eff_first in raw_first_field.split()         # roster first is a middle name in TRACER
    )


# ---------------------------------------------------------------------------
# Step 1 — Load CSV into district-keyed data structure
# ---------------------------------------------------------------------------

def load_races() -> dict:
    """Read the TRACER legislative CSV and build a nested candidate data structure.

    This is the primary data-loading function for legislative (Senate + House)
    races.  It is imported by ingest.py and called during the database build.

    Reads:
        data/tracer_2026_all_districts.csv  (produced by scraper.py)

    Returns:
        A nested dict keyed by chamber → district number → race data:
        {
          "Senate": {
            "3": {
              "label":      "Senate District 3",
              "candidates": [
                {
                  "name":      "SMITH, JOHN A",   # raw TRACER format (LAST, FIRST)
                  "party":     "Democratic",
                  "committee": "Friends of John Smith",
                  "status":    "Active",           # or "Terminated"
                  "raised":    125000.0,           # MonetaryContributions
                  "spent":     87000.0,            # MonetaryExpenditures
                  "coh":       38000.0,            # EndFundsOnHand (cash on hand)
                  "beg":       0.0,                # BegFundsOnHand (beginning balance)
                  "loans":     0.0,                # LoansReceived (self-loans)
                  "vsl":       "Y",                # AcceptedVSL (voluntary spending limit)
                  "incumbent": True,
                },
                ...
              ]
            },
            ...
          },
          "House": { ... }
        }
    """
    races = {"Senate": {}, "House": {}}
    incumbents = load_incumbents()  # loads the Excel roster for incumbent detection

    with open(CSV_FILE, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            chamber = row["Chamber"]        # "Senate" | "House"
            # Normalize district number: "03" → "3" (consistent with GeoJSON)
            dist    = str(int(row["DistrictNumber"]))

            if chamber not in races:
                continue  # skip any unexpected chamber values

            # Initialize the district entry on first encounter
            if dist not in races[chamber]:
                races[chamber][dist] = {
                    "label":      row["DistrictLabel"],
                    "candidates": [],
                }

            # Build candidate dict from CSV row.
            # Float conversion with `or 0` handles empty strings from TRACER.
            races[chamber][dist]["candidates"].append({
                "name":      row["CandName"],
                "party":     row["Party"],
                "committee": row["CommitteeName"],
                "status":    row["CandidateStatus"],
                "raised":    float(row["MonetaryContributions"] or 0),
                "spent":     float(row["MonetaryExpenditures"]  or 0),
                "coh":       float(row["EndFundsOnHand"]        or 0),
                "beg":       float(row["BegFundsOnHand"]        or 0),
                "loans":     float(row["LoansReceived"]         or 0),
                "vsl":       row["AcceptedVSL"],
                "incumbent": is_incumbent(row["CandName"], chamber, dist, incumbents),
            })

    return races


def load_statewide_races() -> dict:
    """Read the TRACER statewide CSV and build the statewide candidate structure.

    Similar to load_races() but for the five Colorado executive offices
    (Governor, Lt. Governor, Secretary of State, State Treasurer, AG).
    Incumbent detection is not performed for statewide races.

    Reads:
        data/tracer_2026_statewide.csv  (produced by: python3 scraper.py --statewide)

    Returns:
        {
          "Governor": {
            "label": "Governor",
            "candidates": [{ name, party, committee, status, raised, spent,
                              coh, loans, vsl, incumbent=False }, ...]
          },
          ...
        }
        Returns {} gracefully if the statewide CSV has not been scraped yet.
    """
    if not STATEWIDE_CSV.exists():
        print("  (no statewide CSV — run: python3 scraper.py --statewide)")
        return {}

    offices: dict = {}
    with open(STATEWIDE_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            office = row["DistrictLabel"]
            if office not in offices:
                offices[office] = {"label": office, "candidates": []}
            offices[office]["candidates"].append({
                "name":      row["CandName"],
                "party":     row["Party"],
                "committee": row["CommitteeName"],
                "status":    row["CandidateStatus"],
                "raised":    float(row["MonetaryContributions"] or 0),
                "spent":     float(row["MonetaryExpenditures"]  or 0),
                "coh":       float(row["EndFundsOnHand"]        or 0),
                "loans":     float(row["LoansReceived"]         or 0),
                "vsl":       row["AcceptedVSL"],
                "incumbent": False,   # statewide executive incumbents not tracked
            })

    total = sum(len(v["candidates"]) for v in offices.values())
    print(f"  Statewide: {len(offices)} offices, {total} candidates")
    return offices


# ---------------------------------------------------------------------------
# Step 2 — Download / cache GeoJSON
# ---------------------------------------------------------------------------

def shapefile_to_geojson(chamber: str, precision: int = 5) -> dict:
    """Download (or load cached) Colorado district boundaries as GeoJSON.

    Downloads the US Census cartographic boundary shapefile for Colorado's
    state legislative districts, converts it to a minimal GeoJSON
    FeatureCollection, and caches the result locally so subsequent runs
    don't re-download.

    The shapefile ZIP contains .shp (geometry), .dbf (attributes), and
    .shx (index) files.  pyshp (the `shapefile` package) reads all three
    from in-memory BytesIO objects — no temporary files needed.

    Coordinate precision is rounded to 5 decimal places (~1 metre accuracy)
    to reduce the JSON file size without visible map degradation.

    Args:
        chamber:   "Senate" or "House"
        precision: Decimal places to round coordinates (default 5)

    Returns:
        A GeoJSON FeatureCollection dict.  Each feature has:
            { "type": "Feature",
              "properties": {"district": "3"},   # district number as string
              "geometry":   { "type": ..., "coordinates": [...] } }
    """
    cache = GEOJSON_CACHE[chamber]
    # Return cached GeoJSON if it exists — avoids re-downloading on every run
    if cache.exists():
        print(f"  Using cached {cache.name}")
        with open(cache, encoding="utf-8") as f:
            return json.load(f)

    url    = SHAPEFILE_URLS[chamber]
    dfield = DISTRICT_FIELD[chamber]  # which DBF attribute holds the district number

    print(f"  Downloading {chamber} shapefile from Census...")
    with urllib.request.urlopen(url, timeout=60) as r:
        zip_bytes = r.read()

    print(f"  Converting shapefile → GeoJSON...")
    features = []

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        # Locate the three component files by extension within the ZIP archive
        names    = zf.namelist()
        shp_name = next(n for n in names if n.endswith(".shp"))
        dbf_name = next(n for n in names if n.endswith(".dbf"))
        shx_name = next(n for n in names if n.endswith(".shx"))

        # Read each component into memory as a BytesIO stream so pyshp can
        # parse them without writing temporary files to disk.
        shp = io.BytesIO(zf.read(shp_name))
        dbf = io.BytesIO(zf.read(dbf_name))
        shx = io.BytesIO(zf.read(shx_name))

        sf = shapefile.Reader(shp=shp, dbf=dbf, shx=shx)
        # sf.fields[0] is the deletion flag — skip it to get the real attribute names
        field_names = [f[0] for f in sf.fields[1:]]

        def round_coords(obj):
            """Recursively round coordinate values to `precision` decimal places.
            Works for any depth of nested lists (Polygon, MultiPolygon, etc.)."""
            if isinstance(obj, list):
                if obj and isinstance(obj[0], (int, float)):
                    # Leaf node: [longitude, latitude]
                    return [round(obj[0], precision), round(obj[1], precision)]
                return [round_coords(c) for c in obj]
            return obj

        for rec, shape in zip(sf.records(), sf.shapes()):
            props = dict(zip(field_names, rec))
            dist_code = props.get(dfield, "")
            try:
                # Normalize "003" or "03" → "3" to match the CSV district numbers
                dist_num = str(int(dist_code))
            except (ValueError, TypeError):
                continue  # skip features with non-numeric district codes

            # __geo_interface__ converts pyshp's shape to a GeoJSON-compatible dict
            geom = shape.__geo_interface__
            features.append({
                "type": "Feature",
                "properties": {"district": dist_num},
                "geometry": {
                    "type":        geom["type"],
                    "coordinates": round_coords(geom["coordinates"]),
                },
            })

    result = {"type": "FeatureCollection", "features": features}

    # Cache for future runs
    with open(cache, "w", encoding="utf-8") as f:
        json.dump(result, f)
    print(f"  Cached → {cache.name}  ({len(features)} polygons)")
    return result


# ---------------------------------------------------------------------------
# Step 3 — City → district map
# ---------------------------------------------------------------------------

def load_places() -> dict:
    """Download (or load cached) Colorado incorporated place centroids.

    Uses the Census TIGER/Line Places shapefile to build a lookup of city
    name → approximate centroid coordinates.  The centroid is estimated as
    the midpoint of the feature's bounding box, which is sufficient accuracy
    for the point-in-polygon district lookup that follows.

    Returns:
        { "Denver": [-104.9903, 39.7392], "Aurora": [...], ... }
        Keys are place NAME strings from the shapefile's DBF attributes.
        Cached to data/co_places.json after the first download.
    """
    if PLACES_CACHE.exists():
        print(f"  Using cached {PLACES_CACHE.name}")
        with open(PLACES_CACHE, encoding="utf-8") as f:
            return json.load(f)

    print("  Downloading Colorado places from Census...")
    with urllib.request.urlopen(PLACES_URL, timeout=60) as r:
        zip_bytes = r.read()

    places = {}
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names    = zf.namelist()
        shp_name = next(n for n in names if n.endswith(".shp"))
        dbf_name = next(n for n in names if n.endswith(".dbf"))
        shx_name = next(n for n in names if n.endswith(".shx"))

        sf = shapefile.Reader(
            shp=io.BytesIO(zf.read(shp_name)),
            dbf=io.BytesIO(zf.read(dbf_name)),
            shx=io.BytesIO(zf.read(shx_name)),
        )
        field_names = [f[0] for f in sf.fields[1:]]

        for rec, shape in zip(sf.records(), sf.shapes()):
            props = dict(zip(field_names, rec))
            name  = props.get("NAME", "").strip()
            if not name:
                continue
            b   = shape.bbox           # [xmin, ymin, xmax, ymax]
            lon = (b[0] + b[2]) / 2
            lat = (b[1] + b[3]) / 2
            places[name] = [lon, lat]

    with open(PLACES_CACHE, "w", encoding="utf-8") as f:
        json.dump(places, f)
    print(f"  Cached → {PLACES_CACHE.name}  ({len(places)} places)")
    return places


def _point_in_polygon(px: float, py: float, feature: dict) -> bool:
    """Test whether a point lies inside a GeoJSON polygon feature.

    Uses the ray-casting algorithm (even-odd rule): counts how many times a
    horizontal ray cast from the test point crosses the polygon boundary.
    An odd number of crossings means the point is inside.

    Only the outer ring of each polygon/multi-polygon is tested.  Holes
    (inner rings) are ignored, which is acceptable here because city
    centroids are unlikely to fall exactly in a lake or park cutout.

    Args:
        px:      Longitude of the test point.
        py:      Latitude of the test point.
        feature: GeoJSON Feature dict with geometry of type Polygon or
                 MultiPolygon.

    Returns:
        True if the point is inside any polygon ring of the feature.
    """
    geom  = feature["geometry"]
    gtype = geom["type"]
    rings = (
        [geom["coordinates"][0]]          if gtype == "Polygon"      else
        [part[0] for part in geom["coordinates"]] if gtype == "MultiPolygon" else
        []
    )
    for ring in rings:
        inside, j = False, len(ring) - 1
        for i, (xi, yi) in enumerate(ring):
            xj, yj = ring[j]
            if ((yi > py) != (yj > py)) and (px < (xj - xi) * (py - yi) / (yj - yi) + xi):
                inside = not inside
            j = i
        if inside:
            return True
    return False


def build_city_map(geojson_senate: dict, geojson_house: dict, places: dict) -> dict:
    """Build a mapping of legislative districts to the cities they contain.

    For every Senate and House district polygon, tests each place centroid
    against the polygon using _point_in_polygon().  Cities whose centroids
    fall inside a district are assigned to it.

    This runs O(districts × places), which is manageable (~200 districts ×
    ~500 places = ~100k checks) and runs in a few seconds.

    Args:
        geojson_senate: FeatureCollection dict from shapefile_to_geojson("Senate")
        geojson_house:  FeatureCollection dict from shapefile_to_geojson("House")
        places:         { city_name: [lon, lat] } from load_places()

    Returns:
        {
          "Senate": { "3": ["Denver", "Glendale", ...], "4": [...], ... },
          "House":  { "6": ["Aurora", ...], ... }
        }
        Districts with no matching place centroids are omitted from the dict.
    """
    city_map    = {"Senate": {}, "House": {}}
    gj_chambers = {"Senate": geojson_senate, "House": geojson_house}

    for chamber, gj in gj_chambers.items():
        for feature in gj["features"]:
            dist   = feature["properties"]["district"]
            cities = sorted(
                name for name, (lon, lat) in places.items()
                if _point_in_polygon(lon, lat, feature)
            )
            if cities:
                city_map[chamber][dist] = cities

    return city_map


# ---------------------------------------------------------------------------
# Step 4 — Generate HTML
# ---------------------------------------------------------------------------

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Colorado 2026 Legislative Fundraising</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
html, body { height: 100%; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; }

/* ── Layout ─────────────────────────────────────────── */
#app { display: flex; flex-direction: column; height: 100vh; }

#topbar {
  display: flex; align-items: center; gap: 12px; flex-wrap: wrap;
  padding: 8px 14px; background: #1e293b; color: #f1f5f9;
  font-size: 13px; flex-shrink: 0; min-height: 50px;
}
#topbar h1 { font-size: 14px; font-weight: 600; color: #e2e8f0; margin-right: 4px; }

#main { display: flex; flex: 1; overflow: hidden; }
#map  { flex: 1; }

/* ── Controls ───────────────────────────────────────── */
.chamber-group { display: flex; background: #334155; border-radius: 6px; overflow: hidden; }
.chamber-btn {
  border: none; padding: 5px 12px; cursor: pointer; font-size: 12px; font-weight: 500;
  color: #94a3b8; background: transparent; transition: all .15s;
}
.chamber-btn.active { background: #3b82f6; color: #fff; }
.chamber-btn:hover:not(.active) { background: #475569; color: #e2e8f0; }

#colorMode {
  border: 1px solid #475569; background: #334155; color: #e2e8f0;
  padding: 5px 8px; border-radius: 6px; font-size: 12px; cursor: pointer;
}

/* ── Legend ─────────────────────────────────────────── */
#legend { display: flex; align-items: center; gap: 6px; margin-left: auto; }
#legend-label { font-size: 11px; color: #94a3b8; }
#legend-bar {
  width: 120px; height: 10px; border-radius: 3px;
  background: linear-gradient(to right, #e02424, #f0f4ff, #1a56db);
}
#legend-ends { display: flex; justify-content: space-between; width: 120px; font-size: 10px; color: #64748b; }

/* ── Sidebar ─────────────────────────────────────────── */
#sidebar {
  width: 0; overflow: hidden; background: #fff;
  border-left: 1px solid #e2e8f0; transition: width .25s ease;
  display: flex; flex-direction: column;
}
#sidebar.open { width: 320px; }

#sidebar-inner { width: 320px; padding: 16px; overflow-y: auto; flex: 1; }

#sidebar-header {
  display: flex; justify-content: space-between; align-items: flex-start;
  margin-bottom: 14px;
}
#sidebar-title { font-size: 15px; font-weight: 700; color: #0f172a; }
#sidebar-close {
  border: none; background: none; cursor: pointer; color: #64748b;
  font-size: 20px; line-height: 1; padding: 0 4px;
}
#sidebar-close:hover { color: #0f172a; }

#sidebar-filters { display: none; gap: 6px; margin-bottom: 12px; flex-wrap: wrap; }
#sidebar.open #sidebar-filters { display: flex; }

.summary-row {
  display: flex; gap: 8px; margin-bottom: 14px;
}
.summary-card {
  flex: 1; background: #f8fafc; border: 1px solid #e2e8f0;
  border-radius: 8px; padding: 8px 10px;
}
.summary-card .label { font-size: 10px; color: #64748b; text-transform: uppercase; letter-spacing: .04em; }
.summary-card .value { font-size: 14px; font-weight: 700; color: #0f172a; margin-top: 2px; }

.section-title {
  font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: .06em;
  color: #64748b; margin: 14px 0 8px;
}

/* Party bar */
.party-bar-wrap { margin-bottom: 14px; }
.party-bar-labels { display: flex; justify-content: space-between; font-size: 11px; margin-bottom: 3px; }
.party-bar-labels .dem { color: #1a56db; font-weight: 600; }
.party-bar-labels .rep { color: #e02424; font-weight: 600; }
.party-bar-track {
  height: 10px; background: #e5e7eb; border-radius: 5px; overflow: hidden; position: relative;
}
.party-bar-fill {
  position: absolute; left: 0; top: 0; height: 100%;
  background: linear-gradient(to right, #1a56db, #93c5fd);
  border-radius: 5px; transition: width .4s ease;
}

/* Candidate cards */
.cand-card {
  position: relative; overflow: hidden;
  border: 1px solid #e2e8f0; border-radius: 8px; padding: 10px 12px; margin-bottom: 8px;
}
.cand-card::after {
  content: ''; position: absolute; top: 0; right: 0;
  width: 0; height: 0; border-style: solid; border-width: 0 22px 22px 0;
}
.cand-card.vsl-yes::after { border-color: transparent #16a34a transparent transparent; }
.cand-card.vsl-no::after  { border-color: transparent #dc2626 transparent transparent; }
.cand-card.inactive {
  background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
  border-color: #e2e8f0; opacity: 0.72;
}
.cand-card.inactive .cand-name { color: #475569; }
.cand-card.inactive .stat-cell .val { color: #475569; }
.cand-name { font-size: 13px; font-weight: 600; color: #0f172a; }
.cand-committee { font-size: 10px; color: #94a3b8; margin-bottom: 3px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.loan-note { font-size: 9px; color: #b45309; font-weight: 600; margin-top: 3px; letter-spacing: .02em; }
.cand-meta { display: flex; align-items: center; gap: 6px; margin: 3px 0 8px; }
.party-badge {
  font-size: 10px; font-weight: 600; padding: 1px 6px; border-radius: 10px;
  text-transform: uppercase; letter-spacing: .04em;
}
.badge-Democratic  { background: #dbeafe; color: #1e40af; }
.badge-Republican  { background: #fee2e2; color: #991b1b; }
.badge-Unaffiliated{ background: #f3f4f6; color: #4b5563; }
.badge-Unknown     { background: #f3f4f6; color: #4b5563; }
.badge-default     { background: #f3f4f6; color: #4b5563; }
.cand-status { font-size: 10px; color: #9ca3af; }

.stat-grid { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 6px; margin-bottom: 8px; }
.stat-cell .lbl { font-size: 9px; color: #9ca3af; text-transform: uppercase; letter-spacing: .04em; }
.stat-cell .val { font-size: 12px; font-weight: 600; color: #0f172a; }

.raised-bar-track {
  height: 6px; background: #e5e7eb; border-radius: 3px; overflow: hidden;
}
.raised-bar-spent { height: 100%; border-radius: 3px; }

/* ── Search ──────────────────────────────────────────── */
#search-wrap { position: relative; }
#search-input {
  border: 1px solid #475569; background: #334155; color: #e2e8f0;
  padding: 5px 10px; border-radius: 6px; font-size: 12px; width: 200px; outline: none;
}
#search-input::placeholder { color: #64748b; }
#search-input:focus { border-color: #3b82f6; }
#search-results {
  display: none; position: absolute; top: calc(100% + 4px); left: 0;
  width: 300px; background: #fff; border: 1px solid #e2e8f0;
  border-radius: 8px; box-shadow: 0 4px 16px rgba(0,0,0,.15);
  z-index: 9999; overflow: hidden; max-height: 340px; overflow-y: auto;
}
#search-results.visible { display: block; }
.search-result {
  padding: 9px 12px; cursor: pointer; border-bottom: 1px solid #f1f5f9;
  display: flex; align-items: flex-start; gap: 8px;
}
.search-result:last-child { border-bottom: none; }
.search-result:hover { background: #f8fafc; }
.sr-badge {
  font-size: 10px; font-weight: 700; padding: 2px 6px; border-radius: 4px;
  background: #334155; color: #e2e8f0; flex-shrink: 0; margin-top: 1px;
}
.sr-main { flex: 1; min-width: 0; }
.sr-district { font-size: 12px; font-weight: 600; color: #0f172a; }
.sr-detail { font-size: 11px; color: #64748b; margin-top: 1px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.sr-empty { padding: 12px; text-align: center; color: #94a3b8; font-size: 12px; }

/* ── Statewide buttons (topbar) ──────────────────────── */
#statewide-divider { width: 1px; height: 24px; background: #475569; flex-shrink: 0; }
#statewide-group { display: flex; gap: 6px; flex-wrap: wrap; }
.statewide-btn {
  border: 1px solid #475569; background: transparent; color: #94a3b8;
  padding: 4px 10px; border-radius: 6px; cursor: pointer;
  font-size: 12px; font-weight: 500; transition: all .15s;
}
.statewide-btn:hover { background: #334155; color: #e2e8f0; }
.statewide-btn.active { background: #7c3aed; border-color: #7c3aed; color: #fff; }

/* ── Filter toggles (statewide panel + sidebar) ──────── */
.filter-toggle {
  border: 1px solid #e2e8f0; background: #fff; color: #64748b;
  padding: 4px 10px; border-radius: 6px; cursor: pointer; font-size: 11px;
  font-weight: 500; transition: all .15s; white-space: nowrap;
}
.filter-toggle:hover { background: #f1f5f9; color: #0f172a; }
.filter-toggle.active { background: #0f172a; border-color: #0f172a; color: #fff; }

/* ── Statewide panel ──────────────────────────────────── */
#statewide-panel {
  flex: 1; display: flex; flex-direction: column; overflow: hidden;
  background: #f8fafc;
}
#statewide-header {
  display: flex; align-items: center; gap: 14px; flex-shrink: 0;
  padding: 10px 20px; background: #fff; border-bottom: 1px solid #e2e8f0;
}
#statewide-back {
  border: 1px solid #e2e8f0; background: #fff; color: #475569;
  padding: 5px 12px; border-radius: 6px; cursor: pointer; font-size: 12px;
  font-weight: 500; transition: all .15s; white-space: nowrap;
}
#statewide-back:hover { background: #f1f5f9; color: #0f172a; }
#statewide-title { font-size: 16px; font-weight: 700; color: #0f172a; flex: 1; }
#statewide-metric {
  border: 1px solid #e2e8f0; background: #fff; color: #334155;
  padding: 5px 8px; border-radius: 6px; font-size: 12px; cursor: pointer;
}
#statewide-chart-wrap { flex: 1; overflow-y: auto; padding: 24px 32px; }
#statewide-chart { display: flex; flex-direction: column; gap: 12px; max-width: 900px; }

/* ── Bar chart rows ───────────────────────────────────── */
.chart-row {
  display: grid; grid-template-columns: 200px 1fr 90px;
  align-items: center; gap: 12px;
}
.chart-label {
  text-align: right; font-size: 13px; color: #1e293b;
  white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}
.chart-sublabel { display: block; font-size: 10px; color: #94a3b8; margin-top: 1px; }
.chart-bar-track { height: 28px; background: #e5e7eb; border-radius: 4px; overflow: hidden; }
.chart-bar-fill { height: 100%; border-radius: 4px; transition: width .4s ease; }
.chart-bar-fill.party-D     { background: #2563eb; }
.chart-bar-fill.party-R     { background: #dc2626; }
.chart-bar-fill.party-other { background: #6366f1; }
.chart-value { font-size: 13px; font-weight: 600; color: #0f172a; white-space: nowrap; }
.chart-inactive { opacity: 0.45; }
.chart-section-header {
  font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: .06em;
  color: #94a3b8; padding: 4px 0 2px; border-bottom: 1px solid #e5e7eb; margin-bottom: 4px;
  grid-column: 1 / -1;
}

/* Leaflet overrides */
.district-label {
  background: transparent; border: none; box-shadow: none;
  font-size: 9px; font-weight: 700; color: #1e293b;
  text-shadow: 0 0 3px #fff, 0 0 3px #fff;
  pointer-events: none;
}
.leaflet-tooltip.district-label::before { display: none; }

/* ── Incumbent badge ──────────────────────────────────── */
.incumbent-badge {
  display: inline-flex; align-items: center; justify-content: center;
  width: 15px; height: 15px; border-radius: 3px;
  background: #7c3aed; color: #fff; font-size: 9px; font-weight: 700;
  margin-left: 5px; vertical-align: middle; flex-shrink: 0; cursor: default;
}

/* ── Topbar icon buttons ────────────────────────────────── */
.icon-btn {
  border: 1px solid #475569; background: transparent; color: #94a3b8;
  padding: 5px 8px; border-radius: 6px; cursor: pointer; font-size: 14px;
  line-height: 1; transition: all .15s;
}
.icon-btn:hover { background: #334155; color: #e2e8f0; }

/* ── Hide VSL indicator when toggled off ──────────────── */
body.vsl-hidden .cand-card::after { display: none; }

/* ── Modals ───────────────────────────────────────────── */
.modal-overlay {
  display: none; position: fixed; inset: 0; background: rgba(0,0,0,.45);
  z-index: 10000; align-items: center; justify-content: center;
}
.modal-overlay.open { display: flex; }
.modal-box {
  background: #fff; border-radius: 10px; padding: 20px 24px;
  max-width: 540px; width: 92%; max-height: 80vh; overflow-y: auto;
  box-shadow: 0 8px 40px rgba(0,0,0,.25);
}
.modal-header {
  display: flex; justify-content: space-between; align-items: center;
  margin-bottom: 14px;
}
.modal-title { font-size: 15px; font-weight: 700; color: #0f172a; }
.modal-close {
  border: none; background: none; cursor: pointer; color: #64748b;
  font-size: 22px; line-height: 1; padding: 0 4px;
}
.modal-close:hover { color: #0f172a; }
.modal-section { margin-bottom: 14px; }
.modal-section h3 {
  font-size: 11px; font-weight: 700; text-transform: uppercase;
  letter-spacing: .06em; color: #64748b; margin-bottom: 6px;
}
.modal-section p, .modal-section li { font-size: 13px; color: #374151; line-height: 1.55; }
.modal-section ul { padding-left: 16px; }
.modal-section li { margin-bottom: 5px; }
.toggle-row {
  display: flex; align-items: flex-start; gap: 10px; cursor: pointer;
  padding: 8px 0; border-bottom: 1px solid #f1f5f9;
}
.toggle-row input[type=checkbox] { margin-top: 2px; cursor: pointer; }
.toggle-row .toggle-label { font-size: 13px; color: #0f172a; font-weight: 500; }
.toggle-row .toggle-desc { font-size: 11px; color: #64748b; margin-top: 2px; }
</style>
</head>
<body>
<div id="app">

  <div id="topbar">
    <h1>Colorado 2026 Legislative Fundraising</h1>

    <div class="chamber-group">
      <button class="chamber-btn active" data-chamber="House">House</button>
      <button class="chamber-btn" data-chamber="Senate">Senate</button>
    </div>

    <div id="statewide-divider"></div>
    <div id="statewide-group"></div>

    <select id="colorMode">
      <option value="raised_margin">Raised Margin (D vs R)</option>
      <option value="coh_margin">Cash on Hand Margin</option>
      <option value="competitiveness">Competitiveness</option>
      <option value="total_raised">Total Raised</option>
      <option value="burn_rate">Burn Rate</option>
      <option value="loan_reliance">Loan Reliance</option>
    </select>

    <div id="search-wrap">
      <input id="search-input" type="text" placeholder="Search candidate or city…" autocomplete="off">
      <div id="search-results"></div>
    </div>

    <button class="icon-btn" id="info-btn" title="How this map works">ℹ</button>
    <button class="icon-btn" id="settings-btn" title="Settings">⚙</button>

    <div id="legend">
      <span id="legend-label">← R leads · D leads →</span>
      <div>
        <div id="legend-bar"></div>
        <div id="legend-ends"><span id="leg-left">R</span><span id="leg-right">D</span></div>
      </div>
    </div>
  </div>

  <div id="main">
    <div id="map"></div>
    <div id="sidebar">
      <div id="sidebar-inner">
        <div id="sidebar-header">
          <div id="sidebar-title"></div>
          <button id="sidebar-close">×</button>
        </div>
        <div id="sidebar-filters">
          <button class="filter-toggle" data-filter="inactive">Hide inactive</button>
          <button class="filter-toggle" data-filter="empty">Hide $0</button>
        </div>
        <div id="sidebar-body"></div>
      </div>
    </div>
  </div>

  <div id="statewide-panel" style="display:none">
    <div id="statewide-header">
      <button id="statewide-back">← Map</button>
      <h2 id="statewide-title"></h2>
      <select id="statewide-metric">
        <option value="raised">Total Raised</option>
        <option value="coh">Cash on Hand</option>
        <option value="spent">Total Spent</option>
        <option value="burn_rate">Burn Rate</option>
        <option value="loan_reliance">Loan Reliance</option>
      </select>
      <button class="filter-toggle" data-filter="inactive">Hide inactive</button>
      <button class="filter-toggle" data-filter="empty">Hide $0</button>
    </div>
    <div id="statewide-chart-wrap">
      <div id="statewide-chart"></div>
    </div>
  </div>

</div>

<!-- ── Info Modal ─────────────────────────────────────── -->
<div class="modal-overlay" id="info-modal">
  <div class="modal-box">
    <div class="modal-header">
      <div class="modal-title">How This Map Works</div>
      <button class="modal-close" id="info-close">×</button>
    </div>
    <div class="modal-section">
      <h3>Data Source</h3>
      <p>All financial data comes from <strong>TRACER</strong>, Colorado's campaign finance reporting system (<em>tracer.sos.colorado.gov</em>), for the 2026 November General Election cycle.</p>
    </div>
    <div class="modal-section">
      <h3>Map Color Modes</h3>
      <ul>
        <li><strong>Raised Margin</strong> — Total contributions raised by Democrats vs. Republicans. Blue = Dem advantage, Red = Rep advantage.</li>
        <li><strong>Cash on Hand Margin</strong> — Same comparison using end-of-period cash available to spend.</li>
        <li><strong>Competitiveness</strong> — How close the fundraising gap is between parties. Brighter amber = tighter race.</li>
        <li><strong>Total Raised</strong> — Combined contributions across all candidates in the district. Darker = more total money.</li>
        <li><strong>Burn Rate</strong> — Ratio of spending to fundraising. Red = spending most of what was raised.</li>
        <li><strong>Loan Reliance</strong> — Share of funds coming from candidate self-loans. Amber = heavy self-funding.</li>
      </ul>
    </div>
    <div class="modal-section">
      <h3>Candidate Card</h3>
      <ul>
        <li><strong>Raised</strong> — Total monetary contributions received this cycle.</li>
        <li><strong>Spent</strong> — Total monetary expenditures this cycle.</li>
        <li><strong>Cash on Hand</strong> — End-of-period available funds (beginning balance + raised − spent).</li>
        <li><strong>Spend bar</strong> — Visual indicator of what portion of raised funds have been spent.</li>
        <li><strong>Loan note</strong> — Candidate has self-funded part of their campaign via personal loans.</li>
        <li><strong>Colored corner triangle</strong> — Green = accepted Volunteer Spending Limits (VSL); Red = declined. VSL is a voluntary cap on spending in exchange for matching public funds eligibility.</li>
        <li><strong>★ purple badge</strong> — Incumbent: candidate currently holds this seat (indicated by non-zero beginning balance from prior cycle).</li>
      </ul>
    </div>
    <div class="modal-section">
      <h3>Sidebar Filters</h3>
      <ul>
        <li><strong>Hide inactive</strong> — Removes terminated and withdrawn candidates from view.</li>
        <li><strong>Hide $0</strong> — Removes candidates who have raised and hold $0.</li>
      </ul>
    </div>
  </div>
</div>

<!-- ── Settings Modal ──────────────────────────────────── -->
<div class="modal-overlay" id="settings-modal">
  <div class="modal-box">
    <div class="modal-header">
      <div class="modal-title">Settings</div>
      <button class="modal-close" id="settings-close">×</button>
    </div>
    <label class="toggle-row">
      <input type="checkbox" id="vsl-toggle" checked>
      <div>
        <div class="toggle-label">Show VSL indicator</div>
        <div class="toggle-desc">Displays a colored corner triangle on candidate cards showing whether they accepted Volunteer Spending Limits.</div>
      </div>
    </label>
  </div>
</div>

<script>
// ── Embedded data ─────────────────────────────────────
const RACES     = __RACES_JSON__;
const STATEWIDE = __STATEWIDE_JSON__;
const GEOJSON   = __GEOJSON_JSON__;
const CITY_MAP  = __CITY_MAP_JSON__;

// ── State ─────────────────────────────────────────────
let activeChamber         = 'House';
let activeMode            = 'raised_margin';
let activeLayer           = null;
let labelLayer            = null;
let selectedDist          = null;
const districtLayerMap    = {};  // 'House:4' → Leaflet layer
let activeStatewide       = null;   // null = map view; string = office name
let activeStatewideMetric = 'raised';
let filterHideInactive    = false;
let filterHideEmpty       = false;

// ── Leaflet setup ─────────────────────────────────────
const map = L.map('map', { zoomControl: true }).setView([38.95, -105.55], 7);

L.tileLayer('https://{s}.basemaps.cartocdn.com/light_nolabels/{z}/{x}/{y}{r}.png', {
  attribution: '© OpenStreetMap © CARTO',
  subdomains: 'abcd', maxZoom: 19
}).addTo(map);

// ── Colour helpers ────────────────────────────────────
function lerpColor(c1, c2, t) {
  t = Math.max(0, Math.min(1, t));
  const h = c => [parseInt(c.slice(1,3),16), parseInt(c.slice(3,5),16), parseInt(c.slice(5,7),16)];
  const [r1,g1,b1] = h(c1), [r2,g2,b2] = h(c2);
  return `rgb(${Math.round(r1+(r2-r1)*t)},${Math.round(g1+(g2-g1)*t)},${Math.round(b1+(b2-b1)*t)})`;
}
function marginColor(m) {
  return m >= 0 ? lerpColor('#f0f4ff','#1a56db', m)
                : lerpColor('#f0f4ff','#e02424',-m);
}
function sum(arr, k) { return arr.reduce((s,c) => s + (c[k]||0), 0); }

// Pre-compute per-chamber max total raised for normalisation
const DISTRICT_MAX = {};
const LOAN_MAX = {};
['Senate','House'].forEach(ch => {
  DISTRICT_MAX[ch] = Math.max(1, ...Object.values(RACES[ch]||{})
    .map(d => sum(d.candidates,'raised')));
  LOAN_MAX[ch] = Math.max(1, ...Object.values(RACES[ch]||{})
    .map(d => sum(d.candidates,'loans')));
});

function districtColor(chamber, distNum, mode) {
  const data = RACES[chamber]?.[String(distNum)];
  if (!data) return '#e2e8f0';   // district not up this cycle

  const cands = data.candidates;
  const dems  = cands.filter(c => c.party === 'Democratic');
  const reps  = cands.filter(c => c.party === 'Republican');

  const Dr = sum(dems,'raised'), Rr = sum(reps,'raised');
  const Dc = sum(dems,'coh'),    Rc = sum(reps,'coh');

  switch (mode) {
    case 'raised_margin':
      if (Dr+Rr === 0) return '#d1d5db';
      return marginColor((Dr-Rr)/(Dr+Rr));

    case 'coh_margin':
      if (Dc+Rc === 0) return '#d1d5db';
      return marginColor((Dc-Rc)/(Dc+Rc));

    case 'competitiveness': {
      if (Dr+Rr === 0) return '#d1d5db';
      const score = 1 - Math.abs((Dr-Rr)/(Dr+Rr));
      return lerpColor('#e5e7eb','#f59e0b', score);
    }
    case 'total_raised': {
      const t = sum(cands,'raised') / DISTRICT_MAX[chamber];
      return lerpColor('#dbeafe','#1e3a8a', t);
    }
    case 'burn_rate': {
      const r = sum(cands,'raised');
      if (r === 0) return '#d1d5db';
      const rate = Math.min(sum(cands,'spent') / r, 1);
      return lerpColor('#d1fae5','#991b1b', rate);
    }
    case 'loan_reliance': {
      const l = sum(cands,'loans');
      if (l === 0) return '#f8fafc';
      return lerpColor('#fef9c3','#b45309', l / LOAN_MAX[chamber]);
    }
  }
  return '#d1d5db';
}

// ── Legend update ─────────────────────────────────────
const LEGEND_CONFIGS = {
  raised_margin:   { bar: 'linear-gradient(to right,#e02424,#f0f4ff,#1a56db)', label:'← R leads · D leads →', l:'R', r:'D' },
  coh_margin:      { bar: 'linear-gradient(to right,#e02424,#f0f4ff,#1a56db)', label:'← R CoH lead · D CoH lead →', l:'R', r:'D' },
  competitiveness: { bar: 'linear-gradient(to right,#e5e7eb,#f59e0b)',          label:'Competitiveness', l:'Safe', r:'Toss-up' },
  total_raised:    { bar: 'linear-gradient(to right,#dbeafe,#1e3a8a)',          label:'Total Raised', l:'$0', r:'Max' },
  burn_rate:       { bar: 'linear-gradient(to right,#d1fae5,#991b1b)',          label:'Burn Rate', l:'Low', r:'High' },
  loan_reliance:   { bar: 'linear-gradient(to right,#fef9c3,#b45309)',          label:'Loan Reliance', l:'None', r:'High' },
};
function updateLegend() {
  const cfg = LEGEND_CONFIGS[activeMode];
  document.getElementById('legend-bar').style.background   = cfg.bar;
  document.getElementById('legend-label').textContent       = cfg.label;
  document.getElementById('leg-left').textContent           = cfg.l;
  document.getElementById('leg-right').textContent          = cfg.r;
}

// ── Polygon centroid helpers (Issue #1: accurate label placement) ──────────
function polygonCentroid(ring) {
  let cx = 0, cy = 0, area = 0;
  for (let i = 0, j = ring.length - 1; i < ring.length; j = i++) {
    const f = ring[i][0] * ring[j][1] - ring[j][0] * ring[i][1];
    cx += (ring[i][0] + ring[j][0]) * f;
    cy += (ring[i][1] + ring[j][1]) * f;
    area += f;
  }
  area /= 2;
  if (Math.abs(area) < 1e-10) {
    return L.latLng(
      ring.reduce((s, p) => s + p[1], 0) / ring.length,
      ring.reduce((s, p) => s + p[0], 0) / ring.length
    );
  }
  return L.latLng(cy / (6 * area), cx / (6 * area));
}

function featureCentroid(feature) {
  const geom = feature.geometry;
  if (geom.type === 'Polygon') {
    return polygonCentroid(geom.coordinates[0]);
  }
  if (geom.type === 'MultiPolygon') {
    let maxArea = 0, bestRing = null;
    for (const poly of geom.coordinates) {
      const ring = poly[0];
      let area = 0;
      for (let i = 0, j = ring.length - 1; i < ring.length; j = i++)
        area += (ring[j][0] + ring[i][0]) * (ring[j][1] - ring[i][1]);
      area = Math.abs(area / 2);
      if (area > maxArea) { maxArea = area; bestRing = ring; }
    }
    return bestRing ? polygonCentroid(bestRing) : null;
  }
  return null;
}

// ── Map layer ─────────────────────────────────────────
function buildLayer(chamber) {
  if (activeLayer)  { map.removeLayer(activeLayer); activeLayer = null; }
  if (labelLayer)   { map.removeLayer(labelLayer);  labelLayer  = null; }

  // Issue #4: only render districts that have candidates this cycle
  const rawGj = GEOJSON[chamber];
  const filteredFeatures = rawGj.features.filter(f =>
    RACES[chamber]?.[f.properties.district]
  );
  const gj = { ...rawGj, features: filteredFeatures };

  activeLayer = L.geoJSON(gj, {
    style: feat => ({
      fillColor:   districtColor(chamber, feat.properties.district, activeMode),
      fillOpacity: 0.75,
      color:       '#475569',
      weight:      0.8,
    }),
    onEachFeature: (feat, layer) => {
      const dist = feat.properties.district;
      districtLayerMap[`${chamber}:${dist}`] = layer;
      layer.on({
        click:     () => openSidebar(chamber, dist),
        mouseover: e  => { e.target.setStyle({ weight: 2, color: '#0f172a' }); },
        mouseout:  e  => { activeLayer.resetStyle(e.target); },
      });
    },
  }).addTo(map);

  // Issue #1: use true polygon centroid for district labels
  labelLayer = L.layerGroup().addTo(map);
  activeLayer.eachLayer(layer => {
    const dist = layer.feature.properties.district;
    const center = featureCentroid(layer.feature) || layer.getBounds().getCenter();
    L.tooltip({ permanent: true, direction: 'center', className: 'district-label', interactive: false })
      .setContent(dist)
      .setLatLng(center)
      .addTo(labelLayer);
  });
}

function refreshColors() {
  if (!activeLayer) return;
  activeLayer.setStyle(feat => ({
    fillColor:   districtColor(activeChamber, feat.properties.district, activeMode),
    fillOpacity: 0.75,
    color:       '#475569',
    weight:      0.8,
  }));
}

// ── Sidebar ───────────────────────────────────────────
function fmt(n) {
  if (n === 0) return '$0';
  if (n >= 1_000_000) return '$' + (n/1_000_000).toFixed(2) + 'M';
  if (n >= 1_000)     return '$' + (n/1_000).toFixed(1) + 'K';
  return '$' + n.toFixed(0);
}

function partyClass(p) {
  return ['Democratic','Republican','Unaffiliated','Unknown'].includes(p) ? `badge-${p}` : 'badge-default';
}

function toTitleCase(s) {
  return s.replace(/\b\w+/g, w =>
    /^(II|III|IV|VI|VII|VIII|IX|JR|SR)$/i.test(w)
      ? w.toUpperCase()
      : w[0].toUpperCase() + w.slice(1).toLowerCase()
  );
}
function fmtName(raw) {
  const comma = raw.indexOf(',');
  if (comma === -1) return toTitleCase(raw);
  return toTitleCase(raw.slice(comma + 1).trim() + ' ' + raw.slice(0, comma).trim());
}

function openSidebar(chamber, distNum) {
  selectedDist = distNum;
  const data   = RACES[chamber]?.[String(distNum)];
  const sidebar = document.getElementById('sidebar');
  const title   = document.getElementById('sidebar-title');
  const body    = document.getElementById('sidebar-body');

  if (!data) {
    title.textContent = `${chamber} District ${distNum}`;
    body.innerHTML    = '<p style="color:#64748b;font-size:13px;margin-top:8px;">No candidates filed for 2026.</p>';
    sidebar.classList.add('open');
    return;
  }

  title.textContent = data.label;

  const cands    = data.candidates;
  const dems     = cands.filter(c => c.party === 'Democratic');
  const reps     = cands.filter(c => c.party === 'Republican');
  const Dr       = sum(dems,'raised'), Rr = sum(reps,'raised');
  const totalR   = sum(cands,'raised');
  const totalS   = sum(cands,'spent');
  const totalCoH = sum(cands,'coh');

  // Party bar pct (D share of D+R only)
  const partyTotal = Dr + Rr;
  const dPct = partyTotal > 0 ? (Dr / partyTotal * 100).toFixed(1) : 50;

  // Filter + sort: Active first, then by raised desc
  let filtered = [...cands];
  if (filterHideInactive) filtered = filtered.filter(c => c.status === 'Active');
  if (filterHideEmpty)    filtered = filtered.filter(c => c.raised > 0 || c.coh > 0);
  const sorted = filtered.sort((a,b) => {
    if (a.status === b.status) return b.raised - a.raised;
    return a.status === 'Active' ? -1 : 1;
  });

  const PARTY_COLOR = { Democratic: '#1a56db', Republican: '#e02424' };
  const candCards = sorted.map(c => {
    const spentPct       = c.raised > 0 ? Math.min(c.spent / c.raised * 100, 100).toFixed(0) : 0;
    const barColor       = PARTY_COLOR[c.party] || '#6366f1';
    const inactive       = c.status !== 'Active';
    const vslClass       = c.vsl === 'Yes' ? 'vsl-yes' : 'vsl-no';
    const incumbentBadge = c.incumbent
      ? '<span class="incumbent-badge" title="Incumbent">★</span>' : '';
    return `
    <div class="cand-card${inactive ? ' inactive' : ''} ${vslClass}">
      <div class="cand-name">${fmtName(c.name)}${incumbentBadge}</div>
      ${c.committee ? `<div class="cand-committee">${c.committee}</div>` : ''}
      <div class="cand-meta">
        <span class="party-badge ${partyClass(c.party)}">${c.party}</span>
        <span class="cand-status">${c.status}</span>
      </div>
      <div class="stat-grid">
        <div class="stat-cell"><div class="lbl">Raised</div><div class="val">${fmt(c.raised)}</div></div>
        <div class="stat-cell"><div class="lbl">Spent</div><div class="val">${fmt(c.spent)}</div></div>
        <div class="stat-cell"><div class="lbl">Cash on Hand</div><div class="val">${fmt(c.coh)}</div></div>
      </div>
      <div class="raised-bar-track">
        <div class="raised-bar-spent" style="width:${spentPct}%;background:${barColor}"></div>
      </div>
      ${c.loans > 0 ? `<div class="loan-note">+ ${fmt(c.loans)} loan</div>` : ''}
    </div>`;
  }).join('');

  body.innerHTML = `
    <div class="summary-row">
      <div class="summary-card"><div class="label">Total Raised</div><div class="value">${fmt(totalR)}</div></div>
      <div class="summary-card"><div class="label">Total Spent</div><div class="value">${fmt(totalS)}</div></div>
      <div class="summary-card"><div class="label">Cash on Hand</div><div class="value">${fmt(totalCoH)}</div></div>
    </div>

    ${partyTotal > 0 ? `
    <div class="party-bar-wrap">
      <div class="party-bar-labels">
        <span class="dem">Dem ${fmt(Dr)}</span>
        <span class="rep">${fmt(Rr)} Rep</span>
      </div>
      <div class="party-bar-track">
        <div class="party-bar-fill" style="width:${dPct}%"></div>
      </div>
    </div>` : ''}

    <div class="section-title">Candidates (${cands.length})${dems.length ? ` <span style="color:#1a56db;font-weight:700">${dems.length}D</span>` : ''}${reps.length ? ` <span style="color:#e02424;font-weight:700">${reps.length}R</span>` : ''}${cands.length - dems.length - reps.length ? ` <span style="color:#9ca3af">${cands.length - dems.length - reps.length} other</span>` : ''}</div>
    ${candCards}
  `;

  sidebar.classList.add('open');
}

document.getElementById('sidebar-close').addEventListener('click', () => {
  document.getElementById('sidebar').classList.remove('open');
  selectedDist = null;
});

// ── Statewide panel ───────────────────────────────────
function showStatewidePanel(office) {
  activeStatewide = office;
  document.querySelectorAll('.chamber-btn').forEach(b => b.classList.remove('active'));
  document.querySelectorAll('.statewide-btn').forEach(b =>
    b.classList.toggle('active', b.dataset.office === office));
  document.getElementById('main').style.display            = 'none';
  document.getElementById('statewide-panel').style.display = 'flex';
  document.getElementById('sidebar').classList.remove('open');
  document.getElementById('colorMode').style.display    = 'none';
  document.getElementById('search-wrap').style.display  = 'none';
  document.getElementById('legend').style.display       = 'none';
  document.getElementById('statewide-title').textContent = office;
  renderStatewideChart(office, activeStatewideMetric);
}

function hideStatewidePanel() {
  activeStatewide = null;
  document.getElementById('main').style.display            = 'flex';
  document.getElementById('statewide-panel').style.display = 'none';
  document.getElementById('colorMode').style.display    = '';
  document.getElementById('search-wrap').style.display  = '';
  document.getElementById('legend').style.display       = '';
  document.querySelectorAll('.statewide-btn').forEach(b => b.classList.remove('active'));
  document.querySelector(`.chamber-btn[data-chamber="${activeChamber}"]`).classList.add('active');
  map.invalidateSize();
}

function getMetricValue(cand, metric) {
  if (metric === 'raised')        return cand.raised;
  if (metric === 'coh')           return cand.coh;
  if (metric === 'spent')         return cand.spent;
  if (metric === 'burn_rate')     return cand.raised > 0 ? cand.spent / cand.raised : 0;
  if (metric === 'loan_reliance') return (cand.raised + cand.loans) > 0
                                    ? cand.loans / (cand.raised + cand.loans) : 0;
  return 0;
}

function renderStatewideChart(office, metric) {
  const data  = STATEWIDE[office];
  if (!data) return;
  const chart    = document.getElementById('statewide-chart');
  const isMoney  = !['burn_rate', 'loan_reliance'].includes(metric);
  let candidates = data.candidates;
  if (filterHideInactive) candidates = candidates.filter(c => c.status === 'Active');
  if (filterHideEmpty)    candidates = candidates.filter(c => c.raised > 0 || c.coh > 0);
  const withVals = candidates.map(c => ({ ...c, metricVal: getMetricValue(c, metric) }));
  const maxVal   = Math.max(1, ...withVals.map(c => c.metricVal));

  const PARTY_ORDER = { Democratic: 0, Republican: 1 };
  const sortFn = (a, b) => {
    const po = (PARTY_ORDER[a.party] ?? 2) - (PARTY_ORDER[b.party] ?? 2);
    return po !== 0 ? po : b.metricVal - a.metricVal;
  };
  const active   = withVals.filter(c => c.status === 'Active').sort(sortFn);
  const inactive = withVals.filter(c => c.status !== 'Active').sort(sortFn);

  function barClass(party) {
    if (party === 'Democratic') return 'party-D';
    if (party === 'Republican') return 'party-R';
    return 'party-other';
  }
  function fmtMetric(val) {
    return isMoney ? fmt(val) : (val * 100).toFixed(1) + '%';
  }
  function renderGroup(arr, headerText) {
    if (!arr.length) return '';
    const header = `<div class="chart-row"><div class="chart-section-header">${headerText}</div></div>`;
    const rows = arr.map(c => {
      const pct  = maxVal > 0 ? (c.metricVal / maxVal * 100).toFixed(1) : 0;
      const cls      = c.status !== 'Active' ? ' chart-inactive' : '';
      const sub      = c.committee ? `<span class="chart-sublabel">${c.committee}</span>` : '';
      const incBadge = c.incumbent ? '<span class="incumbent-badge" title="Incumbent">★</span>' : '';
      return `<div class="chart-row${cls}">
        <div class="chart-label">${fmtName(c.name)}${incBadge}${sub}</div>
        <div class="chart-bar-track">
          <div class="chart-bar-fill ${barClass(c.party)}" style="width:${pct}%"></div>
        </div>
        <div class="chart-value">${fmtMetric(c.metricVal)}</div>
      </div>`;
    }).join('');
    return header + rows;
  }

  chart.innerHTML = renderGroup(active, 'Active Candidates')
                  + renderGroup(inactive, 'Terminated / Withdrawn');
}

(function initStatewideButtons() {
  const swGroup = document.getElementById('statewide-group');
  const offices = Object.keys(STATEWIDE);
  if (!offices.length) {
    document.getElementById('statewide-divider').style.display = 'none';
    return;
  }
  offices.forEach(office => {
    const btn = document.createElement('button');
    btn.className = 'statewide-btn';
    btn.dataset.office = office;
    btn.textContent = office;
    btn.addEventListener('click', () => showStatewidePanel(office));
    swGroup.appendChild(btn);
  });
})();

document.getElementById('statewide-back').addEventListener('click', hideStatewidePanel);
document.getElementById('statewide-metric').addEventListener('change', e => {
  activeStatewideMetric = e.target.value;
  if (activeStatewide) renderStatewideChart(activeStatewide, activeStatewideMetric);
});

document.querySelectorAll('.filter-toggle[data-filter]').forEach(btn => {
  btn.addEventListener('click', () => {
    if (btn.dataset.filter === 'inactive') filterHideInactive = !filterHideInactive;
    else                                   filterHideEmpty     = !filterHideEmpty;
    // Sync all matching buttons
    document.querySelectorAll(`.filter-toggle[data-filter="${btn.dataset.filter}"]`)
      .forEach(b => b.classList.toggle('active', btn.dataset.filter === 'inactive'
                                                 ? filterHideInactive : filterHideEmpty));
    // Re-render wherever applicable
    if (activeStatewide) renderStatewideChart(activeStatewide, activeStatewideMetric);
    if (selectedDist)    openSidebar(activeChamber, selectedDist);
  });
});

// ── Controls ──────────────────────────────────────────
document.querySelectorAll('.chamber-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    if (activeStatewide) hideStatewidePanel();
    if (btn.dataset.chamber === activeChamber) return;
    document.querySelectorAll('.chamber-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    activeChamber = btn.dataset.chamber;
    document.getElementById('sidebar').classList.remove('open');
    selectedDist = null;
    buildLayer(activeChamber);
  });
});

document.getElementById('colorMode').addEventListener('change', e => {
  activeMode = e.target.value;
  updateLegend();
  refreshColors();
});

// ── Search ────────────────────────────────────────────
const searchIndex = [];
['Senate','House'].forEach(ch => {
  Object.entries(RACES[ch]||{}).forEach(([dist, data]) => {
    const cities = CITY_MAP[ch]?.[dist] || [];
    const names  = data.candidates.map(c => c.name);
    const comms  = data.candidates.map(c => c.committee).filter(Boolean);
    searchIndex.push({
      chamber: ch, distNum: dist, label: data.label, cities,
      text: [data.label, ...cities, ...names, ...comms].join(' ').toLowerCase(),
    });
  });
});

function doSearch(q) {
  if (!q.trim()) return [];
  const lq = q.toLowerCase();
  return searchIndex.filter(e => e.text.includes(lq)).slice(0, 8);
}

function renderResults(results, query) {
  const el  = document.getElementById('search-results');
  const lq  = query.toLowerCase();
  if (!results.length) {
    el.innerHTML = '<div class="sr-empty">No results</div>';
    el.classList.add('visible');
    return;
  }
  el.innerHTML = results.map(r => {
    const cands = RACES[r.chamber][r.distNum].candidates;
    const matchedCities = r.cities.filter(c => c.toLowerCase().includes(lq));
    const matchedCands  = cands.filter(c =>
      c.name.toLowerCase().includes(lq) || (c.committee||'').toLowerCase().includes(lq));
    const detail = matchedCities.length
      ? matchedCities.slice(0,3).join(', ')
      : matchedCands.length
        ? matchedCands.slice(0,2).map(c => fmtName(c.name)).join(', ')
        : (r.cities.slice(0,2).join(', ') || cands.slice(0,1).map(c => fmtName(c.name)).join(''));
    return `<div class="search-result" data-chamber="${r.chamber}" data-dist="${r.distNum}">
      <span class="sr-badge">${r.chamber === 'Senate' ? 'S' : 'H'}</span>
      <div class="sr-main">
        <div class="sr-district">${r.label}</div>
        <div class="sr-detail">${detail}</div>
      </div>
    </div>`;
  }).join('');
  el.classList.add('visible');
  el.querySelectorAll('.search-result').forEach(item => {
    item.addEventListener('mousedown', e => {
      e.preventDefault();  // prevent blur firing before click
      goToDistrict(item.dataset.chamber, item.dataset.dist);
    });
  });
}

function goToDistrict(chamber, distNum) {
  // Switch chamber if needed (synchronous — buildLayer runs inline)
  if (chamber !== activeChamber) {
    document.querySelector(`.chamber-btn[data-chamber="${chamber}"]`).click();
  }
  const layer = districtLayerMap[`${chamber}:${distNum}`];
  if (layer) map.fitBounds(layer.getBounds(), { padding: [40, 40] });
  openSidebar(chamber, distNum);
  // Clear search
  document.getElementById('search-input').value = '';
  document.getElementById('search-results').classList.remove('visible');
}

(function () {
  const input   = document.getElementById('search-input');
  const results = document.getElementById('search-results');
  let timer;
  input.addEventListener('input', () => {
    clearTimeout(timer);
    timer = setTimeout(() => {
      const q = input.value;
      if (!q.trim()) { results.classList.remove('visible'); return; }
      renderResults(doSearch(q), q);
    }, 150);
  });
  input.addEventListener('blur', () => {
    setTimeout(() => results.classList.remove('visible'), 150);
  });
  input.addEventListener('focus', () => {
    if (input.value.trim()) renderResults(doSearch(input.value), input.value);
  });
  document.addEventListener('keydown', e => {
    if (e.key === 'Escape') { input.value = ''; results.classList.remove('visible'); input.blur(); }
  });
})();

// ── Modals (Info & Settings) ───────────────────────────
function openModal(id)  { document.getElementById(id).classList.add('open'); }
function closeModal(id) { document.getElementById(id).classList.remove('open'); }

document.getElementById('info-btn').addEventListener('click', () => openModal('info-modal'));
document.getElementById('info-close').addEventListener('click', () => closeModal('info-modal'));
document.getElementById('settings-btn').addEventListener('click', () => openModal('settings-modal'));
document.getElementById('settings-close').addEventListener('click', () => closeModal('settings-modal'));

['info-modal', 'settings-modal'].forEach(id => {
  document.getElementById(id).addEventListener('click', e => {
    if (e.target === e.currentTarget) closeModal(id);
  });
});

document.addEventListener('keydown', e => {
  if (e.key === 'Escape') ['info-modal','settings-modal'].forEach(closeModal);
});

// VSL toggle (Issue #3)
document.getElementById('vsl-toggle').addEventListener('change', e => {
  document.body.classList.toggle('vsl-hidden', !e.target.checked);
});

// ── Init ──────────────────────────────────────────────
// URL param deep-link: /map/?chamber=House&district=44
(function() {
  const p = new URLSearchParams(location.search);
  const ch = p.get('chamber'), dist = p.get('district');
  if (ch && ['Senate','House'].includes(ch)) {
    activeChamber = ch;
    document.querySelectorAll('.chamber-btn').forEach(b =>
      b.classList.toggle('active', b.dataset.chamber === ch));
  }
  if (dist) setTimeout(() => openSidebar(activeChamber, dist), 100);
})();
updateLegend();
buildLayer(activeChamber);
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# Step 5 — Homepage, race pages, and candidate detail pages
# ---------------------------------------------------------------------------

_ROOT_DIR       = Path(__file__).parent
_RACES_DIR      = _ROOT_DIR / "races"
_CANDIDATES_DIR = _ROOT_DIR / "candidates"


def _fmt_dollars(n: float) -> str:
    """Format a dollar amount as a compact human-readable string.

    Thresholds:
        ≥ $1,000,000 → "$1.2M"    (1 decimal, millions suffix)
        ≥ $1,000     → "$12,345"  (comma-separated, no cents)
        < $1,000     → "$45"      (no cents)

    Used as a Jinja2 filter ({{ candidate.raised | fmt_dollars }})
    and directly in the generated HTML map.
    """
    if n >= 1_000_000:
        return f"${n / 1_000_000:.1f}M"
    if n >= 1_000:
        return f"${n:,.0f}"
    return f"${n:.0f}"


def _fmt_name(tracer_name: str) -> str:
    """'LAST, FIRST MIDDLE' → 'First Last'"""
    parts = tracer_name.strip().split(",", 1)
    last = parts[0].strip().title()
    if len(parts) > 1:
        tokens = parts[1].strip().split()
        first = tokens[0].title() if tokens else ""
        return f"{first} {last}".strip() if first else last
    return last


def _slugify(text: str) -> str:
    """Convert a string to a URL-safe slug.

    Steps: lowercase → strip non-word chars → collapse whitespace/underscores
    to hyphens → deduplicate hyphens → trim leading/trailing hyphens.

    Examples:
        "Senate District 3"  → "senate-district-3"
        "Attorney General"   → "attorney-general"
    """
    t = text.lower().strip()
    t = re.sub(r"[^\w\s-]", "", t)
    t = re.sub(r"[\s_]+", "-", t)
    return re.sub(r"-+", "-", t).strip("-")


def _candidate_slug(name: str, context: str) -> str:
    """Generate a unique URL slug for a candidate.

    Combines the candidate's first + last name with the race context to
    create slugs that are unique even when two candidates share a last name.

    Args:
        name:    TRACER name string "LAST, FIRST [MIDDLE]"
        context: "House-44" or "Senate-3" for legislative;
                 office name (e.g. "Governor") for statewide.

    Returns:
        Slug string e.g. "john-smith-house-44" or "jane-doe-governor".
    """
    parts = name.strip().split(",", 1)
    last  = _slugify(parts[0])
    first = _slugify(parts[1].strip().split()[0]) if len(parts) > 1 and parts[1].strip() else ""
    suffix = _slugify(context)
    return f"{first}-{last}-{suffix}" if first else f"{last}-{suffix}"


def _fmt_label(label: str) -> str:
    """Normalize district labels to title case ('HOUSE DISTRICT 44' → 'House District 44')."""
    return label.title()


def _party_badge(party: str) -> str:
    cls = {"Democratic": "badge-D", "Republican": "badge-R"}.get(party, "badge-other")
    return f'<span class="party-badge {cls}">{party}</span>'


def _burn_rate(spent: float, raised: float) -> str:
    """Calculate the percentage of raised funds that have been spent.

    Returns "—" when raised is 0 to avoid division-by-zero.
    Used as a Jinja2 filter: {{ cand.raised | burn_rate(cand.spent) }}
    Note: the filter is registered with args swapped (see app.py).
    """
    return f"{spent / raised * 100:.0f}%" if raised > 0 else "—"


_SHARED_CSS = """\
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
html, body {
  min-height: 100%;
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  background: #f8fafc; color: #0f172a;
}
a { color: #2563eb; text-decoration: none; }
a:hover { text-decoration: underline; }
.site-nav {
  display: flex; align-items: center; gap: 16px; padding: 0 20px;
  background: #1e293b; color: #f1f5f9; height: 50px;
  position: sticky; top: 0; z-index: 100;
}
.nav-brand { font-size: 14px; font-weight: 700; color: #e2e8f0; text-decoration: none; }
.nav-links { display: flex; gap: 16px; margin-left: auto; }
.nav-links a { font-size: 13px; color: #94a3b8; text-decoration: none; }
.nav-links a:hover { color: #e2e8f0; }
.nav-links a.active { color: #fff; font-weight: 600; }
.container { max-width: 900px; margin: 0 auto; padding: 32px 20px; }
.page-title { font-size: 22px; font-weight: 700; color: #0f172a; margin-bottom: 6px; }
.page-subtitle { font-size: 14px; color: #64748b; margin-bottom: 24px; }
.section-label {
  font-size: 11px; font-weight: 700; text-transform: uppercase;
  letter-spacing: .07em; color: #64748b; margin: 24px 0 10px;
}
.party-badge {
  font-size: 10px; font-weight: 700; padding: 2px 8px; border-radius: 10px;
  text-transform: uppercase; letter-spacing: .04em; display: inline-block;
}
.badge-D     { background: #dbeafe; color: #1e40af; }
.badge-R     { background: #fee2e2; color: #991b1b; }
.badge-other { background: #f3f4f6; color: #4b5563; }
.incumbent-badge {
  font-size: 10px; font-weight: 700; background: #ede9fe; color: #6d28d9;
  border-radius: 10px; padding: 2px 8px; display: inline-block;
}
.vsl-badge {
  font-size: 10px; font-weight: 700; border-radius: 10px;
  padding: 2px 8px; display: inline-block;
}
.vsl-yes { background: #dcfce7; color: #15803d; }
.vsl-no  { background: #fee2e2; color: #991b1b; }
.stat-grid {
  display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin: 16px 0;
}
.stat-cell {
  background: #fff; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px 14px;
}
.stat-cell .lbl {
  font-size: 10px; color: #64748b; text-transform: uppercase;
  letter-spacing: .05em; margin-bottom: 4px;
}
.stat-cell .val { font-size: 17px; font-weight: 700; color: #0f172a; }
.party-bar-wrap { margin: 16px 0; }
.party-bar-labels { display: flex; justify-content: space-between; font-size: 12px; margin-bottom: 4px; }
.party-bar-labels .dem { color: #1a56db; font-weight: 700; }
.party-bar-labels .rep { color: #e02424; font-weight: 700; }
.party-bar-track {
  height: 10px; background: #e5e7eb; border-radius: 5px; overflow: hidden; position: relative;
}
.party-bar-fill {
  position: absolute; left: 0; top: 0; height: 100%;
  background: linear-gradient(to right, #1a56db, #93c5fd); border-radius: 5px;
}
.cand-list { display: flex; flex-direction: column; gap: 10px; margin-top: 8px; }
.cand-row {
  display: flex; align-items: center; gap: 12px; justify-content: space-between;
  background: #fff; border: 1px solid #e2e8f0; border-radius: 10px; padding: 12px 16px;
  text-decoration: none; color: inherit; transition: border-color .15s, box-shadow .15s;
}
.cand-row:hover {
  border-color: #93c5fd; box-shadow: 0 2px 8px rgba(37,99,235,.08); text-decoration: none;
}
.cand-row.inactive { opacity: .65; }
.cand-row-left { flex: 1; min-width: 0; }
.cand-row-name { font-size: 14px; font-weight: 600; color: #0f172a; margin-bottom: 4px; }
.cand-row-meta { display: flex; align-items: center; gap: 6px; flex-wrap: wrap; }
.cand-row-committee { font-size: 11px; color: #94a3b8; margin-top: 3px; }
.cand-row-right { display: flex; flex-direction: column; align-items: flex-end; gap: 4px; flex-shrink: 0; }
.cand-row-raised { font-size: 15px; font-weight: 700; color: #0f172a; }
.cand-row-raised-lbl { font-size: 10px; color: #94a3b8; text-transform: uppercase; }
.back-link {
  display: inline-flex; align-items: center; gap: 6px;
  font-size: 13px; color: #64748b; margin-bottom: 20px;
}
.back-link:hover { color: #0f172a; text-decoration: none; }
.city-pills { display: flex; flex-wrap: wrap; gap: 6px; margin-top: 10px; }
.city-pill {
  font-size: 12px; background: #f1f5f9; border: 1px solid #e2e8f0;
  border-radius: 20px; padding: 3px 10px; color: #475569;
}
.btn-row { display: flex; gap: 10px; flex-wrap: wrap; margin: 20px 0; }
.action-btn {
  display: inline-flex; align-items: center; gap: 6px;
  background: #1e293b; color: #f1f5f9;
  padding: 9px 18px; border-radius: 8px; font-size: 13px; font-weight: 500;
  text-decoration: none; transition: background .15s;
}
.action-btn:hover { background: #334155; text-decoration: none; color: #f1f5f9; }
.action-btn.secondary {
  background: #fff; color: #374151; border: 1px solid #d1d5db;
}
.action-btn.secondary:hover { background: #f9fafb; color: #374151; }
.detail-row {
  display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-bottom: 12px;
}
.detail-item .dl {
  font-size: 11px; color: #64748b; text-transform: uppercase;
  letter-spacing: .05em; margin-bottom: 3px;
}
.detail-item .dv { font-size: 14px; color: #0f172a; font-weight: 500; }
.coming-soon-box {
  text-align: center; padding: 40px 20px; background: #fff;
  border: 2px dashed #e2e8f0; border-radius: 12px; color: #94a3b8;
}
.coming-soon-box .cs-icon { font-size: 36px; margin-bottom: 12px; }
.coming-soon-box p { font-size: 14px; margin-bottom: 6px; }
"""


def _nav_html(active: str = "home") -> str:
    items = ""
    for key, url, label in [("home", "/", "Home"), ("map", "/map/", "Map")]:
        cls = ' class="active"' if key == active else ""
        items += f'<a href="{url}"{cls}>{label}</a>'
    return (
        '<nav class="site-nav">'
        '<a href="/" class="nav-brand">Colorado 2026 Legislative Fundraising</a>'
        f'<div class="nav-links">{items}</div>'
        '</nav>'
    )


# ── Homepage ──────────────────────────────────────────────────────────────

_HOMEPAGE_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Colorado 2026 Legislative Fundraising</title>
<style>
__SHARED_CSS__
.hero {
  background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
  padding: 56px 20px 48px; text-align: center; color: #f1f5f9;
}
.hero h1 {
  font-size: 28px; font-weight: 800; color: #fff;
  margin-bottom: 8px; letter-spacing: -.02em;
}
.hero p { font-size: 15px; color: #94a3b8; margin-bottom: 32px; }
.search-outer { position: relative; max-width: 540px; margin: 0 auto; }
.search-bar {
  width: 100%; padding: 14px 20px; font-size: 15px;
  border: 2px solid #334155; background: #1e293b; color: #f1f5f9;
  border-radius: 12px; outline: none; transition: border-color .2s;
}
.search-bar::placeholder { color: #475569; }
.search-bar:focus { border-color: #3b82f6; }
.search-hint { margin-top: 10px; font-size: 12px; color: #475569; }
#hp-results {
  display: none; position: absolute; top: calc(100% + 6px); left: 0;
  width: 100%; background: #fff; border: 1px solid #e2e8f0;
  border-radius: 12px; box-shadow: 0 8px 32px rgba(0,0,0,.18);
  z-index: 200; overflow: hidden; max-height: 420px; overflow-y: auto; text-align: left;
}
#hp-results.visible { display: block; }
.sr-section-head {
  padding: 8px 14px 4px; font-size: 10px; font-weight: 700;
  text-transform: uppercase; letter-spacing: .07em; color: #94a3b8;
  border-top: 1px solid #f1f5f9;
}
.sr-section-head:first-child { border-top: none; }
.sr-item {
  display: flex; align-items: center; gap: 10px; padding: 10px 14px;
  cursor: pointer; border-bottom: 1px solid #f8fafc;
  text-decoration: none; color: inherit;
}
.sr-item:last-child { border-bottom: none; }
.sr-item:hover { background: #f8fafc; }
.sr-type {
  font-size: 10px; font-weight: 700; padding: 2px 7px;
  border-radius: 4px; flex-shrink: 0;
}
.sr-type-race { background: #334155; color: #e2e8f0; }
.sr-type-cand { background: #dbeafe; color: #1e40af; }
.sr-body { flex: 1; min-width: 0; }
.sr-title { font-size: 13px; font-weight: 600; color: #0f172a; }
.sr-sub {
  font-size: 11px; color: #64748b; margin-top: 2px;
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
}
.sr-empty { padding: 16px; text-align: center; color: #94a3b8; font-size: 13px; }
.cards-section {
  max-width: 900px; margin: 40px auto; padding: 0 20px;
  display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px;
}
.feature-card {
  background: #fff; border: 1px solid #e2e8f0; border-radius: 14px;
  padding: 24px 20px; transition: all .2s;
  text-decoration: none; color: inherit; display: flex; flex-direction: column;
  cursor: pointer;
}
.feature-card:hover {
  border-color: #93c5fd; box-shadow: 0 4px 20px rgba(37,99,235,.1); text-decoration: none;
}
.feature-card.disabled { cursor: default; opacity: .65; }
.feature-card.disabled:hover { border-color: #e2e8f0; box-shadow: none; }
.card-icon { font-size: 28px; margin-bottom: 12px; }
.card-title { font-size: 15px; font-weight: 700; color: #0f172a; margin-bottom: 6px; }
.card-desc { font-size: 13px; color: #64748b; line-height: 1.5; flex: 1; }
.card-badge {
  display: inline-block; margin-top: 12px; font-size: 11px; font-weight: 600;
  padding: 3px 10px; border-radius: 20px; background: #f1f5f9; color: #64748b;
}
.card-cta { margin-top: 16px; font-size: 13px; color: #3b82f6; font-weight: 600; }
</style>
</head>
<body>
__NAV__
<div class="hero">
  <h1>Colorado 2026 Campaign Finance</h1>
  <p>Explore fundraising data for all legislative and statewide races.</p>
  <div class="search-outer">
    <input id="hp-search" class="search-bar" type="text"
           placeholder="Search races, candidates, cities&#8230;" autocomplete="off">
    <div id="hp-results"></div>
  </div>
  <p class="search-hint">Try &ldquo;House District 44&rdquo;, &ldquo;Denver&rdquo;, or a candidate name</p>
</div>
<div class="cards-section">
  <div class="feature-card" onclick="focusSearch()">
    <div class="card-icon">&#128499;&#65039;</div>
    <div class="card-title">Find a Race</div>
    <div class="card-desc">Browse all House and Senate districts and statewide offices. See every candidate and their fundraising totals.</div>
    <div class="card-cta">Search above &#8593;</div>
  </div>
  <div class="feature-card" onclick="focusSearch()">
    <div class="card-icon">&#128100;</div>
    <div class="card-title">Find a Candidate</div>
    <div class="card-desc">Look up any candidate by name or committee to see their full financial breakdown.</div>
    <div class="card-cta">Search above &#8593;</div>
  </div>
  <div class="feature-card disabled">
    <div class="card-icon">&#128176;</div>
    <div class="card-title">Find a Donor</div>
    <div class="card-desc">Search contributions by donor name to see all donations made across Colorado campaigns.</div>
    <span class="card-badge">Coming Soon</span>
  </div>
  <a class="feature-card" href="/map/">
    <div class="card-icon">&#128506;&#65039;</div>
    <div class="card-title">Browse the Map</div>
    <div class="card-desc">Explore campaign finance data geographically &mdash; click any district to see all candidates.</div>
    <div class="card-cta">Open Map &rarr;</div>
  </a>
</div>
<script>
const RACES    = __RACES_JSON__;
const STATEWIDE = __STATEWIDE_JSON__;
const CITY_MAP = __CITY_MAP_JSON__;

function slugify(t) {
  return t.toLowerCase().trim()
    .replace(/[^\w\s-]/g,'').replace(/[\s_]+/g,'-')
    .replace(/-+/g,'-').replace(/^-|-$/g,'');
}
function fmtName(s) {
  const p = s.trim().split(',');
  const last = p[0].trim().replace(/\b\w/g, c => c.toUpperCase());
  if (p.length > 1) {
    const first = p[1].trim().split(/\s+/)[0].replace(/\b\w/g, c => c.toUpperCase());
    return first + ' ' + last;
  }
  return last;
}
function candSlug(name, ctx) {
  const p = name.trim().split(',');
  const last = slugify(p[0]);
  const first = p.length > 1 ? slugify(p[1].trim().split(/\s+/)[0]) : '';
  const suf = slugify(ctx);
  return first ? first + '-' + last + '-' + suf : last + '-' + suf;
}

const idx = [];
['Senate','House'].forEach(ch => {
  Object.entries(RACES[ch] || {}).forEach(([dist, data]) => {
    const cities = CITY_MAP[ch]?.[dist] || [];
    const names  = data.candidates.map(c => c.name);
    const comms  = data.candidates.map(c => c.committee).filter(Boolean);
    idx.push({
      type:'race', label:data.label,
      url: '/races/' + ch.toLowerCase() + '-' + dist + '/',
      sub: cities.slice(0,3).join(', ') || (data.candidates.length + ' candidates'),
      text: [data.label, ...cities, ...names, ...comms].join(' ').toLowerCase(),
    });
    data.candidates.forEach(c => {
      idx.push({
        type:'candidate', label:fmtName(c.name), party:c.party,
        url: '/candidates/' + candSlug(c.name, ch + '-' + dist) + '/',
        sub: c.party + ' \u00b7 ' + data.label,
        text: [c.name, c.committee||'', data.label, ...cities].join(' ').toLowerCase(),
      });
    });
  });
});
Object.entries(STATEWIDE).forEach(([office, data]) => {
  idx.push({
    type:'race', label:office,
    url: '/races/' + slugify(office) + '/',
    sub: data.candidates.length + ' candidates \u00b7 Statewide',
    text: [office, ...data.candidates.map(c => c.name)].join(' ').toLowerCase(),
  });
  data.candidates.forEach(c => {
    idx.push({
      type:'candidate', label:fmtName(c.name), party:c.party,
      url: '/candidates/' + candSlug(c.name, office) + '/',
      sub: c.party + ' \u00b7 ' + office,
      text: [c.name, c.committee||'', office].join(' ').toLowerCase(),
    });
  });
});

function doSearch(q) {
  if (!q.trim()) return [];
  const lq = q.toLowerCase();
  return idx.filter(e => e.text.includes(lq)).slice(0, 14);
}

const inp = document.getElementById('hp-search');
const box = document.getElementById('hp-results');
let debounce;

inp.addEventListener('input', () => {
  clearTimeout(debounce);
  debounce = setTimeout(() => render(doSearch(inp.value)), 150);
});
inp.addEventListener('keydown', e => {
  if (e.key === 'Escape') {
    inp.value = ''; box.innerHTML = ''; box.classList.remove('visible');
  }
});
document.addEventListener('click', e => {
  if (!e.target.closest('.search-outer')) box.classList.remove('visible');
});

function render(results) {
  if (!results.length) {
    box.innerHTML = inp.value ? '<div class="sr-empty">No results found</div>' : '';
    box.classList.toggle('visible', !!inp.value);
    return;
  }
  const races = results.filter(r => r.type === 'race');
  const cands = results.filter(r => r.type === 'candidate');
  let html = '';
  if (races.length) {
    html += '<div class="sr-section-head">Races</div>';
    races.forEach(r => {
      html += '<a class="sr-item" href="' + r.url + '">'
        + '<span class="sr-type sr-type-race">Race</span>'
        + '<div class="sr-body"><div class="sr-title">' + r.label + '</div>'
        + '<div class="sr-sub">' + r.sub + '</div></div></a>';
    });
  }
  if (cands.length) {
    html += '<div class="sr-section-head">Candidates</div>';
    cands.forEach(r => {
      const bc = r.party === 'Democratic' ? 'badge-D'
               : r.party === 'Republican' ? 'badge-R' : 'badge-other';
      html += '<a class="sr-item" href="' + r.url + '">'
        + '<span class="sr-type sr-type-cand">Candidate</span>'
        + '<div class="sr-body"><div class="sr-title">' + r.label
        + ' <span class="party-badge ' + bc + '">' + r.party + '</span></div>'
        + '<div class="sr-sub">' + r.sub + '</div></div></a>';
    });
  }
  box.innerHTML = html;
  box.classList.add('visible');
}

function focusSearch() {
  inp.focus();
  inp.scrollIntoView({behavior:'smooth', block:'center'});
}
</script>
</body>
</html>
"""


def generate_homepage(races: dict, statewide: dict, city_map: dict) -> None:
    """Generate the root index.html homepage with unified search."""
    races_json     = json.dumps(races,     separators=(',', ':'))
    statewide_json = json.dumps(statewide, separators=(',', ':'))
    city_map_json  = json.dumps(city_map,  separators=(',', ':'))

    html = _HOMEPAGE_TEMPLATE
    html = html.replace('__SHARED_CSS__',     _SHARED_CSS)
    html = html.replace('__NAV__',            _nav_html("home"))
    html = html.replace('__RACES_JSON__',     races_json)
    html = html.replace('__STATEWIDE_JSON__', statewide_json)
    html = html.replace('__CITY_MAP_JSON__',  city_map_json)

    out = _ROOT_DIR / "index.html"
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  \u2713 Homepage \u2192 {out.name}")


# ── Race pages ────────────────────────────────────────────────────────────

_RACE_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__PAGE_TITLE__ &mdash; Colorado 2026</title>
<style>__SHARED_CSS__</style>
</head>
<body>
__NAV__
<div class="container">
  <a class="back-link" href="/">&larr; Back to Search</a>
  <h1 class="page-title">__PAGE_TITLE__</h1>
  <p class="page-subtitle">__PAGE_SUBTITLE__</p>
  __CONTENT__
</div>
</body>
</html>
"""


def _race_page_content(label: str, chamber: str, dist: str,
                       candidates: list, cities: list) -> str:
    """Build the inner HTML content block for a race page."""
    dems = [c for c in candidates if c["party"] == "Democratic"]
    reps = [c for c in candidates if c["party"] == "Republican"]
    dr   = sum(c["raised"] for c in dems)
    rr   = sum(c["raised"] for c in reps)
    total_raised = sum(c["raised"] for c in candidates)
    total_coh    = sum(c["coh"]    for c in candidates)

    # Map link
    if chamber in ("Senate", "House"):
        map_url = f"/map/?chamber={chamber}&district={dist}"
        map_btn = f'<a class="action-btn" href="{map_url}">\U0001f5fa\ufe0f View on Map</a>'
    else:
        map_btn = '<a class="action-btn secondary" href="/map/">\U0001f5fa\ufe0f Open Map</a>'

    # Summary stats
    stats = (
        '<div class="stat-grid">'
        f'<div class="stat-cell"><div class="lbl">Total Raised</div>'
        f'<div class="val">{_fmt_dollars(total_raised)}</div></div>'
        f'<div class="stat-cell"><div class="lbl">Cash on Hand</div>'
        f'<div class="val">{_fmt_dollars(total_coh)}</div></div>'
        f'<div class="stat-cell"><div class="lbl">Candidates</div>'
        f'<div class="val">{len(candidates)}</div></div>'
        '</div>'
    )

    # Party fundraising bar
    party_total = dr + rr
    d_pct = round(dr / party_total * 100, 1) if party_total > 0 else 50
    party_bar = (
        '<div class="party-bar-wrap">'
        '<div class="party-bar-labels">'
        f'<span class="dem">Dem {_fmt_dollars(dr)}</span>'
        f'<span class="rep">Rep {_fmt_dollars(rr)}</span>'
        '</div>'
        '<div class="party-bar-track">'
        f'<div class="party-bar-fill" style="width:{d_pct}%"></div>'
        '</div></div>'
    ) if party_total > 0 else ""

    # City pills (legislative districts only)
    city_section = ""
    if cities:
        pills = "".join(f'<span class="city-pill">{c}</span>' for c in sorted(cities))
        city_section = (
            '<div class="section-label">Communities in this District</div>'
            f'<div class="city-pills">{pills}</div>'
        )

    # Candidate list sorted by raised descending
    cand_items = ""
    for c in sorted(candidates, key=lambda x: -x["raised"]):
        ctx = f"{chamber}-{dist}" if chamber in ("Senate", "House") else chamber
        slug       = _candidate_slug(c["name"], ctx)
        badge      = _party_badge(c["party"])
        inc_html   = '<span class="incumbent-badge">\u2605 Incumbent</span>' if c.get("incumbent") else ""
        vsl_cls    = "vsl-yes" if c.get("vsl") == "Yes" else "vsl-no"
        vsl_lbl    = "VSL \u2713" if c.get("vsl") == "Yes" else "VSL \u2717"
        inactive   = " inactive" if c.get("status") not in (None, "Active") else ""
        comm       = c.get("committee", "") or ""
        comm_html  = f'<div class="cand-row-committee">{comm}</div>' if comm else ""
        cand_items += (
            f'<a class="cand-row{inactive}" href="/candidates/{slug}/">'
            f'<div class="cand-row-left">'
            f'<div class="cand-row-name">{_fmt_name(c["name"])}</div>'
            f'<div class="cand-row-meta">{badge}{inc_html}'
            f'<span class="vsl-badge {vsl_cls}">{vsl_lbl}</span></div>'
            f'{comm_html}'
            f'</div>'
            f'<div class="cand-row-right">'
            f'<div class="cand-row-raised">{_fmt_dollars(c["raised"])}</div>'
            f'<div class="cand-row-raised-lbl">Raised</div>'
            f'</div>'
            f'</a>'
        )

    cand_section = (
        '<div class="section-label">Candidates</div>'
        f'<div class="cand-list">{cand_items}</div>'
    ) if cand_items else (
        '<p style="color:#64748b;font-size:14px;margin-top:16px;">'
        'No candidates on file for 2026.</p>'
    )

    return (
        f'<div class="btn-row">{map_btn}</div>'
        f'{stats}{party_bar}{city_section}{cand_section}'
    )


def generate_race_pages(races: dict, statewide: dict, city_map: dict) -> None:
    """Generate one HTML page per district and statewide office."""
    _RACES_DIR.mkdir(parents=True, exist_ok=True)
    count = 0

    for chamber in ("Senate", "House"):
        for dist, data in races.get(chamber, {}).items():
            slug     = f"{chamber.lower()}-{dist}"
            cities   = city_map.get(chamber, {}).get(dist, [])
            page_dir = _RACES_DIR / slug
            page_dir.mkdir(parents=True, exist_ok=True)

            label    = _fmt_label(data["label"])
            content  = _race_page_content(label, chamber, dist, data["candidates"], cities)
            subtitle = f"{chamber} District \u00b7 {len(data['candidates'])} candidates"
            if cities:
                preview = ", ".join(cities[:3])
                subtitle += f" \u00b7 {preview}" + (f" +{len(cities)-3} more" if len(cities) > 3 else "")

            html = _RACE_TEMPLATE
            html = html.replace('__SHARED_CSS__',    _SHARED_CSS)
            html = html.replace('__NAV__',           _nav_html())
            html = html.replace('__PAGE_TITLE__',    label)
            html = html.replace('__PAGE_SUBTITLE__', subtitle)
            html = html.replace('__CONTENT__',       content)

            with open(page_dir / "index.html", "w", encoding="utf-8") as f:
                f.write(html)
            count += 1

    for office, data in statewide.items():
        slug     = _slugify(office)
        page_dir = _RACES_DIR / slug
        page_dir.mkdir(parents=True, exist_ok=True)

        content  = _race_page_content(office, office, office, data["candidates"], [])
        subtitle = f"Statewide \u00b7 {len(data['candidates'])} candidates"

        html = _RACE_TEMPLATE
        html = html.replace('__SHARED_CSS__',    _SHARED_CSS)
        html = html.replace('__NAV__',           _nav_html())
        html = html.replace('__PAGE_TITLE__',    office)
        html = html.replace('__PAGE_SUBTITLE__', subtitle)
        html = html.replace('__CONTENT__',       content)

        with open(page_dir / "index.html", "w", encoding="utf-8") as f:
            f.write(html)
        count += 1

    print(f"  \u2713 Race pages \u2192 {_RACES_DIR.name}/  ({count} pages)")


# ── Candidate pages ───────────────────────────────────────────────────────

_CANDIDATE_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__PAGE_TITLE__ &mdash; Colorado 2026</title>
<style>__SHARED_CSS__</style>
</head>
<body>
__NAV__
<div class="container">
  <a class="back-link" href="__RACE_URL__">&larr; Back to __RACE_LABEL__</a>
  __CONTENT__
</div>
</body>
</html>
"""


def _candidate_page_content(cand: dict, chamber: str, dist_or_office: str,
                             race_label: str, race_url: str) -> str:
    name      = _fmt_name(cand["name"])
    badge     = _party_badge(cand["party"])
    inc_html  = '<span class="incumbent-badge">\u2605 Incumbent</span>' if cand.get("incumbent") else ""
    status    = cand.get("status", "Active") or "Active"
    vsl       = cand.get("vsl", "")
    committee = cand.get("committee", "") or ""
    raised    = cand.get("raised", 0.0)
    spent     = cand.get("spent",  0.0)
    coh       = cand.get("coh",    0.0)
    loans     = cand.get("loans",  0.0)
    burn      = _burn_rate(spent, raised)
    loan_pct  = f"{loans / (raised + loans) * 100:.0f}%" if (raised + loans) > 0 else "0%"
    vsl_cls   = "vsl-yes" if vsl == "Yes" else "vsl-no"
    vsl_text  = "Accepted VSL" if vsl == "Yes" else ("Declined VSL" if vsl == "No" else "VSL Unknown")

    # TRACER search link via committee name
    tracer_q   = (committee or cand["name"]).replace(" ", "+")
    tracer_url = f"https://tracer.sos.colorado.gov/PublicSite/SearchPages/CommitteeSearch.aspx?name={tracer_q}"

    header = (
        f'<h1 class="page-title">{name}</h1>'
        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;flex-wrap:wrap;">'
        f'{badge}{inc_html}'
        f'<span class="vsl-badge {vsl_cls}">{vsl_text}</span>'
        f'</div>'
    )
    if committee:
        header += f'<p style="font-size:13px;color:#64748b;margin-bottom:4px;">{committee}</p>'
    if status not in (None, "Active"):
        header += f'<p style="font-size:12px;color:#dc2626;font-weight:600;margin-bottom:16px;">Status: {status}</p>'
    else:
        header += '<div style="margin-bottom:16px;"></div>'

    fmt_race_label = _fmt_label(race_label)
    details = (
        '<div class="detail-row">'
        f'<div class="detail-item"><div class="dl">Running For</div>'
        f'<div class="dv"><a href="{race_url}">{fmt_race_label}</a></div></div>'
        f'<div class="detail-item"><div class="dl">Filing Status</div>'
        f'<div class="dv">{status}</div></div>'
        '</div>'
    )

    stats = (
        '<div class="section-label">Fundraising</div>'
        '<div class="stat-grid">'
        f'<div class="stat-cell"><div class="lbl">Raised</div>'
        f'<div class="val">{_fmt_dollars(raised)}</div></div>'
        f'<div class="stat-cell"><div class="lbl">Spent</div>'
        f'<div class="val">{_fmt_dollars(spent)}</div></div>'
        f'<div class="stat-cell"><div class="lbl">Cash on Hand</div>'
        f'<div class="val">{_fmt_dollars(coh)}</div></div>'
        f'<div class="stat-cell"><div class="lbl">Loans</div>'
        f'<div class="val">{_fmt_dollars(loans)}</div></div>'
        f'<div class="stat-cell"><div class="lbl">Burn Rate</div>'
        f'<div class="val">{burn}</div></div>'
        f'<div class="stat-cell"><div class="lbl">Loan Reliance</div>'
        f'<div class="val">{loan_pct}</div></div>'
        '</div>'
    )

    buttons = (
        '<div class="btn-row">'
        f'<a class="action-btn secondary" href="{tracer_url}" target="_blank" rel="noopener">'
        '\u2197 View on TRACER</a>'
        '</div>'
    )

    placeholder = (
        '<div class="section-label">Additional Information</div>'
        '<div class="coming-soon-box">'
        '<div class="cs-icon">\U0001f517</div>'
        '<p style="font-weight:600;color:#475569;">Web links &amp; contact info</p>'
        '<p>Phone, website, and social media links will appear here when available.</p>'
        '</div>'
    )

    return header + details + stats + buttons + placeholder


def generate_candidate_pages(races: dict, statewide: dict) -> None:
    """Generate one HTML page per candidate."""
    _CANDIDATES_DIR.mkdir(parents=True, exist_ok=True)
    count = 0

    for chamber in ("Senate", "House"):
        for dist, data in races.get(chamber, {}).items():
            race_slug = f"{chamber.lower()}-{dist}"
            race_url  = f"/races/{race_slug}/"
            for cand in data["candidates"]:
                slug     = _candidate_slug(cand["name"], f"{chamber}-{dist}")
                page_dir = _CANDIDATES_DIR / slug
                page_dir.mkdir(parents=True, exist_ok=True)

                content = _candidate_page_content(
                    cand, chamber, dist, data["label"], race_url
                )
                html = _CANDIDATE_TEMPLATE
                html = html.replace('__SHARED_CSS__',  _SHARED_CSS)
                html = html.replace('__NAV__',         _nav_html())
                html = html.replace('__PAGE_TITLE__',  _fmt_name(cand["name"]))
                html = html.replace('__RACE_URL__',    race_url)
                html = html.replace('__RACE_LABEL__',  _fmt_label(data["label"]))
                html = html.replace('__CONTENT__',     content)

                with open(page_dir / "index.html", "w", encoding="utf-8") as f:
                    f.write(html)
                count += 1

    for office, data in statewide.items():
        office_slug = _slugify(office)
        race_url    = f"/races/{office_slug}/"
        for cand in data["candidates"]:
            slug     = _candidate_slug(cand["name"], office)
            page_dir = _CANDIDATES_DIR / slug
            page_dir.mkdir(parents=True, exist_ok=True)

            content = _candidate_page_content(
                cand, office, office, office, race_url
            )
            html = _CANDIDATE_TEMPLATE
            html = html.replace('__SHARED_CSS__',  _SHARED_CSS)
            html = html.replace('__NAV__',         _nav_html())
            html = html.replace('__PAGE_TITLE__',  _fmt_name(cand["name"]))
            html = html.replace('__RACE_URL__',    race_url)
            html = html.replace('__RACE_LABEL__',  office)
            html = html.replace('__CONTENT__',     content)

            with open(page_dir / "index.html", "w", encoding="utf-8") as f:
                f.write(html)
            count += 1

    print(f"  \u2713 Candidate pages \u2192 {_CANDIDATES_DIR.name}/  ({count} pages)")


def generate_html(races: dict, statewide: dict, geojson_senate: dict,
                  geojson_house: dict, city_map: dict) -> str:
    races_json     = json.dumps(races,      separators=(',', ':'))
    statewide_json = json.dumps(statewide,  separators=(',', ':'))
    geojson_json   = json.dumps({"Senate": geojson_senate, "House": geojson_house},
                                separators=(',', ':'))
    city_map_json  = json.dumps(city_map,   separators=(',', ':'))

    html = HTML_TEMPLATE
    html = html.replace('__RACES_JSON__',     races_json)
    html = html.replace('__STATEWIDE_JSON__', statewide_json)
    html = html.replace('__GEOJSON_JSON__',   geojson_json)
    html = html.replace('__CITY_MAP_JSON__',  city_map_json)
    return html


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    MAP_DIR.mkdir(parents=True, exist_ok=True)

    print("Loading candidate data...")
    races = load_races()
    senate_count = sum(len(v['candidates']) for v in races['Senate'].values())
    house_count  = sum(len(v['candidates']) for v in races['House'].values())
    print(f"  Senate: {len(races['Senate'])} districts, {senate_count} candidates")
    print(f"  House:  {len(races['House'])} districts, {house_count} candidates")

    print("\nLoading statewide data...")
    statewide = load_statewide_races()

    print("\nLoading GeoJSON...")
    gj_senate = shapefile_to_geojson("Senate")
    gj_house  = shapefile_to_geojson("House")
    print(f"  Senate: {len(gj_senate['features'])} district polygons")
    print(f"  House:  {len(gj_house['features'])} district polygons")

    print("\nLoading Colorado places...")
    places = load_places()

    print("\nBuilding city → district map...")
    city_map = build_city_map(gj_senate, gj_house, places)
    s_mapped = sum(1 for v in city_map["Senate"].values() if v)
    h_mapped = sum(1 for v in city_map["House"].values()  if v)
    print(f"  Senate: {s_mapped} districts with cities")
    print(f"  House:  {h_mapped} districts with cities")

    print(f"\nGenerating {OUTPUT_HTML}...")
    html = generate_html(races, statewide, gj_senate, gj_house, city_map)

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)

    size_kb = OUTPUT_HTML.stat().st_size // 1024
    print(f"  \u2713 Map \u2192 {OUTPUT_HTML} ({size_kb} KB)")

    print("\nGenerating homepage and static pages...")
    generate_homepage(races, statewide, city_map)
    generate_race_pages(races, statewide, city_map)
    generate_candidate_pages(races, statewide)

    print(f"\n\u2713 Done")
    print(f"\nOpen in browser:  open index.html")
    print(f"Or serve locally: python3 -m http.server 8000")


if __name__ == "__main__":
    main()
